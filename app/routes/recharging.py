# app/routes/recharging.py
from app import db 
from flask import Blueprint, render_template, session, jsonify, request, flash, redirect, url_for, send_file, Response
from io import StringIO
from sqlalchemy.orm import joinedload
from app.models import *
from app.routes.logs import add_log
from flask_login import login_required
from app.routes.auth import user_login_required
import pandas as pd
import os, zipfile, subprocess
from datetime import datetime
import calendar
import re
import zipfile
from io import BytesIO
import csv
import io
from xero_python.accounting import Invoice, LineItem, Attachment,LineItemTracking
from flask_login import current_user
from app.celery_tasks import upload_recharge_invoices_xero_task
from app import celery
from app.xero import get_last_invoice_number
import xlsxwriter
from openpyxl import load_workbook

from app.celery_tasks import process_recharging_task

# Define the blueprint
recharging_bp = Blueprint('recharging', __name__, url_prefix='/recharging')

@recharging_bp.route('/task_status/<task_id>', methods=['GET'])
@user_login_required
def celery_task_status(task_id):
    task = celery.AsyncResult(task_id)
    response = {
        'state': task.state,
        'status': str(task.info) if task.info else None
    }

    if task.state == 'SUCCESS':
        # Clear session data related to CSV contents
        session.pop('purchase_csv', None)
        session.pop('breakdown_csv', None)
        session.pop('sales_invoices_csv', None)

    if task.state == 'FAILURE':
        # Optional: If you want more detailed error information in case of failure
        response['error'] = str(task.info) if task.info else "Unknown error"

    return jsonify(response)


def convert_parentheses_to_negative(value):
    # Remove commas
    value = value.replace(',', '')
    # Handle values in parentheses
    if re.match(r'^\(.*\)$', value):
        value = '-' + value[1:-1]
    return value


def combine_last_two_columns(df):
    # Define the possible column name pairs
    column_pairs = [
        ('Southern', 'Northern'),
        ('DOMINOS', 'ALL-GDK-COSTA-GYM'),
        ('tracking_group1', 'tracking_group2'), 
        ('DOMINOS', 'OTHER')
    ]
    
    # Iterate through the possible column name pairs
    for col1, col2 in column_pairs:
        if col1 in df.columns and col2 in df.columns:
            # Combine the two columns into 'Tracking Code'
            df['Tracking Code'] = df.apply(
                lambda row: row[col1] if pd.notna(row[col1]) and row[col1] != '' else row[col2],
                axis=1
            )
            # Drop the original columns
            df = df.drop(columns=[col1, col2])
            break  # Stop once we've handled one pair of columns
    
    return df


@recharging_bp.route('/run', methods=['GET'])
@user_login_required
def recharging_home():
    return render_template('recharging.html', datetime=datetime)


@recharging_bp.route('/get_last_invoice_number', methods=['POST'])
@user_login_required
def get_last_invoice_number_route():
    try:
        # Get the user and selected month/year from the POST request
        user_id = session.get('user_id')
        user = User.query.get(user_id)
        selected_month = int(request.json['selected_month'])
        selected_year = int(request.json['selected_year'])

        # Fetch the last invoice number for the selected month and year
        last_invoice_data = get_last_invoice_number(user, selected_month, selected_year)

        # Check if the function returned an error status
        if last_invoice_data.get('status') == 'error':
            return jsonify({"status": "error", "message": last_invoice_data.get('message')})
        
        # Success case, return the highest invoice number
        return jsonify({"status": "success", "last_invoice_number": last_invoice_data.get('highest_invoice_number')})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})
    


@recharging_bp.route('/run', methods=['POST'])
@user_login_required
def run_recharging():
    try:
        print(f"Form data received: {request.form}")
        user_id = session.get('user_id')
        print(f"user_id from session: {user_id}")

        if not user_id:
            print("Error: No user_id found in session.")
            return jsonify({"status": "error", "message": "User not logged in."}), 400

        selected_month = int(request.form['selected_month'])
        selected_year = int(request.form['selected_year'])
        last_invoice_number = int(request.form['last_invoice_number']) + 1

        task_type = 'recharging'
        print("Checking for active task...")

        active_task = TaskStatus.query.filter_by(user_id=user_id, task_type=task_type, status='in_progress').first()
        print(f"Query successful: active_task={active_task}")

        if active_task:
            print("Task in progress")
            return jsonify({"status": "error", "message": "A recharging task is already in progress."}), 400

        # Trigger Celery task
        try:
            task = process_recharging_task.apply_async(args=[user_id, selected_month, selected_year, last_invoice_number])
            print(f"Task successfully queued with ID: {task.id}")
        except Exception as e:
            print(f"Error queuing task: {e}")
            return jsonify({"status": "error", "message": "Failed to queue task."}), 500

        # Save task details to the database
        try:
            new_task = TaskStatus(
                task_id=task.id,
                user_id=user_id,
                task_type=task_type,
                status='in_progress'
            )
            db.session.add(new_task)
            db.session.commit()
            print(f"TaskStatus entry created for task_id: {task.id}")
        except Exception as e:
            print(f"Error saving TaskStatus to database: {e}")
            return jsonify({"status": "error", "message": "Failed to save task to database."}), 500

        return jsonify({"status": "success", "message": "Processing started", "task_id": task.id})
    except Exception as e:
        error_message = f"Error: {str(e)}"
        print(error_message)
        return jsonify({"status": "error", "message": str(e)}), 500


@recharging_bp.route('/download/<task_id>', methods=['GET'])
def download_file(task_id):
    task_result = process_recharging_task.AsyncResult(task_id)

    # Fetch task from the database
    task_status = TaskStatus.query.filter_by(task_id=task_id).first()

    if not task_status:
        return jsonify({"status": "error", "message": "Task not found"}), 404

    if task_result.state == 'SUCCESS':
        file_path = task_result.result.get('file_path')  # This should now point to `static/temp_files/`

        if os.path.exists(file_path):
            # Mark task as completed if not already done
            if task_status.status != 'completed':
                task_status.status = 'completed'
                db.session.commit()

            # Send the file to the user
            return send_file(
                file_path,
                as_attachment=True,
                download_name=os.path.basename(file_path),  # Suggest the correct file name
                mimetype='application/zip'
            )
        else:
            task_status.status = 'failed'
            task_status.result = "File not found"
            db.session.commit()
            return jsonify({"status": "error", "message": "File not found"}), 404
    elif task_result.state == 'FAILURE':
        task_status.status = 'failed'
        task_status.result = str(task_result.result)
        db.session.commit()
        return jsonify({"status": "error", "message": f"Task failed: {task_status.result}"}), 500
    elif task_result.state == 'PENDING':
        return jsonify({"status": "pending", "message": "Task is still processing"}), 202
    else:
        return jsonify({"status": "error", "message": "Unknown error occurred during task execution"}), 500


@recharging_bp.route('/task-status/<task_id>', methods=['GET'])
def get_task_status(task_id):
    task_status = TaskStatus.query.filter_by(task_id=task_id).first()
    if not task_status:
        return jsonify({"status": "no_task", "message": "No task found."})
    return jsonify({"status": task_status.status, "result": task_status.result or "No result available."})


@recharging_bp.route('/current-task-status/<task_type>', methods=['GET'])
def get_current_task_status(task_type):
    user_id = session.get('user_id')

    # Get the latest task that is still running or recently finished
    task_status = (
        TaskStatus.query
        .filter_by(user_id=user_id, task_type=task_type)
        .filter(TaskStatus.status.in_(['in_progress', 'completed', 'failed']))
        .order_by(TaskStatus.created_at.desc())
        .first()
    )

    if not task_status:
        return jsonify({"status": "no_task", "message": "No task is currently running."})

    return jsonify({
        "status": task_status.status,
        "task_id": task_status.task_id,
        "result": task_status.result or "No result available."
    })


# # Route to handle the 'run macros' functionality (now recharging process)
# @recharging_bp.route('/run', methods=['POST'])
# @user_login_required
# def run_recharging():
#     if request.method == 'POST': 
#         try:
#             user_id = session.get('user_id')
#             user = User.query.get(user_id)
#             current_company = user.company_name
#             current_username = user.username
#             # Log when the process starts
#             add_log("Processing Macro button pressed", log_type="general")

#             selected_month = int(request.form['selected_month'])
#             selected_year = int(request.form['selected_year'])
#             lastInvoiceNumber = int(request.form['last_invoice_number']) + 1  # Now automatically 
#             #lastInvoiceNumber = manual set

#             selected_month_str = f"{selected_year}-{selected_month:02d}"
#             selected_month_year_str_filenames = f"{datetime(1900, selected_month, 1).strftime('%B')} {selected_year}"

#             add_log(f"Selected month: {selected_month_str}", log_type="general")

#             # Query the database for account transactions for the logged-in user
#             transactions = AccountTransaction.query.filter_by(user_id=session['user_id']).all()
            

#             # Example of handling an error case
#             if not transactions:
#                 add_log("No data found for the current user.", log_type="error")
#                 return jsonify({"status": "error", "message":"No data found for the current user. Please check the logs."}), 400

#             add_log(f"Found {len(transactions)} transactions for the user.", log_type="general")

#             # Convert the transaction records to a pandas DataFrame
#             transaction_data = [
#                 {
#                     'Date': t.date,
#                     'Source': t.source,
#                     'Contact': t.contact,
#                     'Description': t.description,
#                     'Reference': t.reference,
#                     'Debit': t.debit,
#                     'Credit': t.credit,
#                     'Gross': t.gross,
#                     'Net': t.net,
#                     'VAT': t.vat,
#                     'Account Code': t.account_code,
#                     'Account': t.account,
#                     'tracking_group1': t.tracking_group1,
#                     'tracking_group2': t.tracking_group2
#                 } for t in transactions
#             ]
#             df_unfiltered = pd.DataFrame(transaction_data)

#             # Calculate the sum of the 'Net' column
#             net_sum = df_unfiltered['Net'].sum()

#             # Print the sum of the 'Net' column
#             print(f"Sum of 'Net' column: {net_sum}")

#             if df_unfiltered.empty:
#                 add_log("No data found in DataFrame after converting transactions.", log_type="error")
#                 return jsonify({"status": "error", "message": "No data found in data.csv."}), 400

#             # Filter by the selected month and year

#             # Step 1: Try to parse the 'Date' with the first format ('%Y-%m-%d')
#             df_unfiltered['Date'] = pd.to_datetime(df_unfiltered['Date'], format='%Y-%m-%d', errors='coerce')

#             # Step 2: For rows where the 'Date' is NaT (failed to parse), try the second format ('%d %b %Y')
#             df_unfiltered['Date'] = df_unfiltered['Date'].fillna(pd.to_datetime(df_unfiltered['Date'], format='%d %b %Y', errors='coerce'))
            

#             df  = df_unfiltered[(df_unfiltered['Date'].dt.month == selected_month) &
#                                         (df_unfiltered['Date'].dt.year == selected_year)]


#             if df.empty:
#                 add_log(f"No transactions found for {selected_month_str}.", log_type="error")
#                 return jsonify({"status": "error", "message": f"No transactions found for {selected_month_str}."}), 400
#             else:
#                 add_log(f"Found {len(df)} transactions for {selected_month_str}.", log_type="general")


#             # Further processing with logs
#             modified_df = combine_last_two_columns(df)
#             columns_to_drop = ['Debit', 'Credit', 'Gross', 'VAT']
#             modified_df = modified_df.drop(columns=columns_to_drop, errors='ignore')
#             modified_df['Account Code Per Business'] = ''
#             modified_df['Account Code Per Business Description'] = ''
#             modified_df['Net'] = modified_df['Net'].astype(float)

#             add_log("DataFrame modified successfully. Dropped unnecessary columns and added business account codes.", log_type="general")

#             # Get all account codes per DMS and map them to business account codes
#             account_codes_per_dms = (
#                 db.session.query(
#                     AccountCodesPerDMS.account_code_per_dms,
#                     AccountCodesPerBusiness.account_code_per_business,
#                     AccountCodesPerDMS.descriptor_per_dms,
#                     AccountCodesPerBusiness.descriptor_per_business
#                 )
#                 .join(AccountCodesPerBusiness, AccountCodesPerDMS.business_id == AccountCodesPerBusiness.id)
#                 .filter(AccountCodesPerDMS.user_id == session['user_id'])
#                 .all()
#             )

#             add_log(f"Fetched {len(account_codes_per_dms)} account codes per DMS.", log_type="general")

#             # Query the database for companies associated with the user
#             companies = Company.query.filter_by(user_id=user_id).all()

#             # Extract company names and codes into a list of dictionaries
#             data = [
#                 {
#                     'company_name': company.company_name,
#                     'company_code': company.company_code
#                 }
#                 for company in companies
#             ]

#             # Create a pandas DataFrame from the data
#             companies_df = pd.DataFrame(data)

#             add_log(f"Fetched {len(companies_df)} companies.", log_type="general")

#             # Convert the query result to a DataFrame
#             account_codes_per_dms_data = [
#                 {
#                     'Account Code Per DMS': code_per_dms[0],
#                     'Account Code Per Business': code_per_dms[1],
#                     'Descriptor Per DMS': code_per_dms[2],
#                     'Descriptor Per Business': code_per_dms[3]
#                 } for code_per_dms in account_codes_per_dms
#             ]
#             account_codes_per_dms_df = pd.DataFrame(account_codes_per_dms_data)

#             # Ensure columns are integers
#             account_codes_per_dms_df['Account Code Per DMS'] = account_codes_per_dms_df['Account Code Per DMS'].astype(int)
#             account_codes_per_dms_df['Account Code Per Business'] = account_codes_per_dms_df['Account Code Per Business'].astype(int)

#             # Further processing (Tracking Codes, Group Codes, etc.)
#             add_log("Tracking codes and group codes processing started.", log_type="general")

#             tracking_codes = TrackingCode.query.filter_by(user_id=session['user_id']).with_entities(TrackingCode.tracking_code).all()
#             tracking_codes = [code[0] for code in tracking_codes if code]

#             group_tracking_codes = GroupTrackingCode.query.filter_by(user_id=session['user_id']).with_entities(GroupTrackingCode.group_code).all()
#             group_tracking_codes = [code[0] for code in group_tracking_codes]

#             group_tracking_code_mappings = GroupTrackingCode.query.filter_by(user_id=session['user_id']).options(joinedload(GroupTrackingCode.tracking_codes)).all()

#             group_tracking_code_map = {}
#             for group in group_tracking_code_mappings:
#                 group_code = group.group_code
#                 tracking_codes_associated = [tc.tracking_code for tc in group.tracking_codes]
#                 group_tracking_code_map[group_code] = tracking_codes_associated


#             # Process each row of modified_df
#             processed_rows = []
#             for _, row in modified_df.iterrows():
#                 tracking_code = row['Tracking Code']
#                 net_value = row['Net']
#                 if tracking_code in tracking_codes:
#                     company_code = tracking_code.split(' -')[0]
#                     row['Company Code'] = company_code
#                     processed_rows.append(row)
#                 elif tracking_code in group_tracking_codes:
#                     assigned_codes = group_tracking_code_map[tracking_code]
#                     num_assigned_codes = len(assigned_codes)
#                     divided_net_value = round(net_value / num_assigned_codes, 2)
#                     total_rounded_value = divided_net_value * (num_assigned_codes - 1)
#                     remainder = round(net_value - total_rounded_value, 2)
#                     for i, code in enumerate(assigned_codes):
#                         new_row = row.copy()
#                         if i == num_assigned_codes - 1:
#                             new_row['Net'] = remainder
#                         else:
#                             new_row['Net'] = divided_net_value
#                         new_row['Tracking Code'] = code
#                         company_code = code.split(' -')[0]
#                         new_row['Company Code'] = company_code
#                         new_row['Description'] = f"{new_row['Description']} - Total Invoice Amount = {net_value}, Tracking Code = {tracking_code}"
#                         processed_rows.append(new_row)
#                 else:
#                     add_log(f"Error: Tracking code {tracking_code} not found in any tracking codes.", log_type="error")
#                     return jsonify({"status": "error", "message": f"Tracking code {tracking_code} not found in any tracking codes."}), 400

#             add_log(f"Processed {len(processed_rows)} rows of tracking data.", log_type="general")

#             final_df = pd.DataFrame(processed_rows)

#             # Match account codes with business codes
#             for index, row in final_df.iterrows():
#                 account_code = row['Account Code']
#                 matching_row = account_codes_per_dms_df[account_codes_per_dms_df['Account Code Per DMS'] == account_code]
#                 if matching_row.empty:
#                     add_log(f"No matching account code per business found for {account_code}.", log_type="error")
#                     return jsonify({"status": "error", "message":f"No matching account code per business found for {account_code}"}), 400
#                 else:
#                     account_code_per_business = matching_row.iloc[0]['Account Code Per Business']
#                     descriptor_per_business = matching_row.iloc[0]['Descriptor Per Business']
#                     final_df.at[index, 'Account Code Per Business'] = account_code_per_business
#                     final_df.at[index, 'Account Code Per Business Description'] = descriptor_per_business

#             add_log("Account codes matched successfully.", log_type="general")

#             # Generate final files, zip them, and return to the user
#             add_log("Generating invoices and zipping them.", log_type="general")

            

#             # Create a new DataFrame for the sales invoice
#             sales_invoice_df = pd.DataFrame(columns=['ContactName','EmailAdress','POAddressLine1','POAddressLine2','POAddressLine3','POAddressLine4','POCity','PORegion','POPostalCode','POCountry', 'InvoiceNumber', 'Company Code', 'InvoiceDate', 'DueDate','Total','InventoryItemCode', 'Description', 'Quantiy', 'UnitAmount','Discount','AccountCode', 'TaxType'])
            
#             starting_invoice_number = lastInvoiceNumber
#             row_count = 0

#             # Get the current date
#             current_date = datetime.now().strftime("%Y-%m-%d")
#             current_month = datetime.now().strftime("%B")
#             current_year = datetime.now().strftime("%Y")

#             # Calculate the last day of the selected month and year
#             last_day_of_month = calendar.monthrange(selected_year, selected_month)[1]

#             # Create the invoice date as the last day of the selected month and year
#             invoice_date = datetime(selected_year, selected_month, last_day_of_month).strftime("%d/%m/%Y")

#             selected_month_name = calendar.month_name[selected_month]


#             # Process each row in the modified DataFrame
#             for _, row in final_df.iterrows():
#                 company_code = row['Company Code']
#                 net_amount = row['Net']

#                 # Find the company name from companies.csv
#                 # Get the company name where the company_code matches
#                 matching_company = companies_df.loc[companies_df['company_code'] == company_code, 'company_name']

#                 # Check if there are any matches
#                 if not matching_company.empty:
#                     company_name = matching_company.values[0]
#                 else:
#                     company_name = None  # Set company_name to None if no match is found
#                     add_log(f"Error: No company found for company code '{company_code}'", log_type="error")


#                 # Check if the company name already exists in the invoice DataFrame
#                 if not sales_invoice_df[(sales_invoice_df['ContactName'] == company_name) & (sales_invoice_df['Company Code'] == company_code)].empty:
#                     # Add the net amount to the existing net amount
#                     sales_invoice_df.loc[(sales_invoice_df['ContactName'] == company_name) & (sales_invoice_df['Company Code'] == company_code), 'UnitAmount'] +=net_amount
#                 else:
#                     #  Add a new row
#                     new_row = pd.DataFrame({
#                         'ContactName': [company_name],
#                         'InvoiceNumber': [str(current_username)+' - ' + str(starting_invoice_number + row_count)],
#                         'Quantity': ['1'],
#                         'TaxType':['20% (VAT on Income)'],
#                         'Description': ['HO and Management cost recharges for the month of ' + str(selected_month_name) + ' ' + str(selected_year)],
#                         'Company Code': [company_code],
#                         'InvoiceDate': [invoice_date],
#                         'DueDate': [invoice_date],
#                         'UnitAmount': [net_amount],
#                         'AccountCode': ['4000']
#                     })
#                     row_count = row_count + 1
#                     sales_invoice_df = pd.concat([sales_invoice_df, new_row], ignore_index=True)
            



#             # Ensure all values in the 'Net' column are rounded to two decimal places
#             sales_invoice_df['UnitAmount'] = sales_invoice_df['UnitAmount'].round(2)


#             #st.write("Sales Invoice Data:")
#             #st.dataframe(sales_invoice_df)

#             # Convert sales_invoice_df to CSV
#             csv_sales = sales_invoice_df.to_csv(index=False).encode('utf-8')

            
#         # Button to produce purchase invoices
#         #if st.button("Produce Purchase Invoices"):


#             # Create a new DataFrame for the purchase invoice
#             purchase_invoice_df = pd.DataFrame(columns=['ContactName','EmailAddress','POAddressLine1','POAddressLine2','POAddressLine3','POAddressLine4','POCity','PORegion','POPostalCode','POCountry','InvoiceNumber', 'InvoiceDate', 'DueDate', 'Total', 'InventoryItemCode', 'Description', 'Quantity', 'UnitAmount', 'AccountCode', 'TaxType', 'TaxAmount', 'TrackingName1', 'TrackingOption1', 'TrackingName2', 'TrackingOption2', 'Currency', 'Company Code', 'Company Name'])

#             # Process each row in the final DataFrame
#             for _, row in final_df.iterrows():
#                 company_code = row['Company Code']
#                 net_amount = row['Net']
#                 #tracking_code = row['Tracking Code'].split('-')[1].strip() 
#                 tracking_code = '-'.join(row['Tracking Code'].split('-')[1:]).strip()

#                 account_code = row['Account Code Per Business']  # Assuming 'Account Code' is a column in your DataFrame


#                 # Find the company name from companies.csv
#                 company_name = companies_df.loc[companies_df['company_code'] == company_code, 'company_name'].values[0]

#                 # Check if the company code and account code already exist in the purchase invoice DataFrame
#                 existing_row = purchase_invoice_df[(purchase_invoice_df['Company Code'] == company_code) & (purchase_invoice_df['AccountCode'] == account_code) & (purchase_invoice_df['TrackingOption1'] == tracking_code) ]
#                 if not existing_row.empty:
#                     # Add the net amount to the existing net amount
#                     purchase_invoice_df.loc[(purchase_invoice_df['Company Code'] == company_code) & (purchase_invoice_df['AccountCode'] == account_code)  & (purchase_invoice_df['TrackingOption1'] == tracking_code) , 'UnitAmount'] += net_amount
#                 else:
#                     # Find the row in the sales_invoice_df with the matching company name
#                     matching_row = sales_invoice_df[sales_invoice_df['ContactName'] == company_name]

#                     # Retrieve the invoice number from the matching row
#                     if not matching_row.empty:
#                         invoice_number = matching_row.iloc[0]['InvoiceNumber']
#                     else:
#                         invoice_number = None  # or handle the case where no match is found
                        
                    
#                     # Add a new row
#                     new_row = pd.DataFrame({
#                         'ContactName': [current_company],
#                         'InvoiceDate': [invoice_date],
#                         'DueDate': [invoice_date],
#                         'Quantity': ['1'],
#                         'TaxType':['20% (VAT on Expenses)'],
#                         'TrackingName1': ['Store'],
#                         'InvoiceNumber': [invoice_number],
#                         'Description': ['HO and Management cost recharges for the month of ' + str(selected_month_name) + ' ' + str(selected_year)],
#                         'TrackingOption1': [tracking_code],
#                         'Company Name': [company_name],
#                         'Company Code': [company_code],
#                         'AccountCode': [account_code],
#                         'UnitAmount': [net_amount]
#                     })
#                     purchase_invoice_df = pd.concat([purchase_invoice_df, new_row], ignore_index=True)
                

#             # Ensure all values in the 'Net' column are rounded to two decimal places
#             purchase_invoice_df['UnitAmount'] = purchase_invoice_df['UnitAmount'].round(2)


#             #st.write("Purchase Invoice Data:")
#             #st.dataframe(purchase_invoice_df)
#             # Create directories if they don't exist
#             # 1. Create purchase_invoices.zip in memory
#             purchase_invoices_zip_stream = BytesIO()
#             with zipfile.ZipFile(purchase_invoices_zip_stream, 'w', zipfile.ZIP_DEFLATED) as purchase_zip:
#                 for company in purchase_invoice_df['Company Name'].unique():
#                     company_df = purchase_invoice_df[purchase_invoice_df['Company Name'] == company]
#                     # Get the invoice number from the first row of the company data
#                     invoice_number = company_df.iloc[0]['InvoiceNumber']
#                     # Add the invoice number at the beginning of the filename
#                     company_filename = f"{invoice_number}-{company}-{selected_month_year_str_filenames}.csv"
#                     csv_content = company_df.to_csv(index=False)
#                     purchase_zip.writestr(company_filename, csv_content)
#             purchase_invoices_zip_stream.seek(0)
#             zip_data1 = purchase_invoices_zip_stream.read()

#             # 2. Create breakdown_invoices.zip in memory
#             breakdown_invoices_zip_stream = BytesIO()
#             with zipfile.ZipFile(breakdown_invoices_zip_stream, 'w', zipfile.ZIP_DEFLATED) as breakdown_zip:
#                 for company_code in final_df['Company Code'].unique():
#                     # Filter the dataframe for the current company
#                     company_df = final_df[final_df['Company Code'] == company_code]
                    
#                     # Get the company name where the company_code matches
#                     matching_company_name = companies_df.loc[companies_df['company_code'] == company_code, 'company_name'].values

#                     # Check if matching_company_name is not empty and extract the first value
#                     if len(matching_company_name) > 0:
#                         company_name_str = matching_company_name[0]
#                     else:
#                         company_name_str = "Unknown Company"

#                     # Sort the transactions by 'Account Code per Business'
#                     company_df_sorted = company_df.sort_values(by='Account Code Per Business')

#                     # Initialize an empty DataFrame for breakdown with totals
#                     breakdown_with_totals = pd.DataFrame()

#                     # Loop through each 'Account Code per Business' and group data
#                     for code, group in company_df_sorted.groupby('Account Code Per Business'):
                        
#                         # Sort each group by 'Net' in descending order
#                         group = group.sort_values(by='Net', ascending=False)
            
#                         # Calculate the total for 'Net' column, rounded to 2 decimal places
#                         total_net = round(group['Net'].sum(), 2)

#                         # Add a 'Total Net' column, setting the total value only on the last row of the group
#                         group['Total Net'] = [None] * (len(group) - 1) + [total_net]

#                         # Concatenate the group with breakdown_with_totals
#                         breakdown_with_totals = pd.concat([breakdown_with_totals, group], ignore_index=True)

#                         # Add an empty row for separation between different account codes
#                         empty_row = pd.Series([None] * len(breakdown_with_totals.columns), index=breakdown_with_totals.columns)
#                         breakdown_with_totals = pd.concat([breakdown_with_totals, pd.DataFrame([empty_row])], ignore_index=True)

#                     # Check and print the column names for debugging
#                     print("Columns after grouping and processing:", breakdown_with_totals.columns)

#                     # Ensure 'Total Net' is one of the columns
#                     if 'Total Net' in breakdown_with_totals.columns and 'Account Code Per Business' in breakdown_with_totals.columns:
#                         # Create a summary table for account codes and their totals (Net column only)
#                         summary_df = breakdown_with_totals[['Account Code Per Business', 'Total Net']].dropna(subset=['Total Net']).reset_index(drop=True)
#                     else:
#                         # If the columns are not present, raise an error
#                         raise ValueError("Required columns 'Account Code Per Business' or 'Total Net' are missing from the DataFrame")

#                     # Align the lengths of both DataFrames for concatenation
#                     max_len = max(len(breakdown_with_totals), len(summary_df))
#                     breakdown_with_totals.reset_index(drop=True, inplace=True)
#                     summary_df.reset_index(drop=True, inplace=True)

#                     # Align lengths for concatenation
#                     breakdown_with_totals = pd.concat([breakdown_with_totals, pd.DataFrame(index=range(max_len))], axis=1)
#                     summary_df = pd.concat([summary_df, pd.DataFrame(index=range(max_len))], axis=1)

#                     # Concatenate the breakdown with totals and the summary table side by side
#                     final_company_df = pd.concat([breakdown_with_totals, summary_df], axis=1)

#                     # Create the filename using the company name and code
#                     company_filename = f"Breakdown for {company_name_str} ({company_code}) - {selected_month_year_str_filenames}.csv"

#                     # Write the CSV content to a string
#                     csv_content = final_company_df.to_csv(index=False)

#                     # Write this string to the zip file
#                     breakdown_zip.writestr(company_filename, csv_content)

#             breakdown_invoices_zip_stream.seek(0)
#             zip_data2 = breakdown_invoices_zip_stream.read()

#             # Convert company and tracking codes to strings
#             final_df['Company Code'] = final_df['Company Code'].astype(str)
#             final_df['Account Code Per Business'] = final_df['Account Code Per Business'].astype(str)
#             final_df['Tracking Code'] = final_df['Tracking Code'].astype(str)
#             final_df['Account Code'] = final_df['Account Code'].astype(str)

#             # Join final_df with companies_df to include company names
#             final_df = final_df.merge(companies_df, left_on='Company Code', right_on='company_code', how='left')
#             final_df.drop(columns=['company_code'], inplace=True)

#             # Prepare the data for Full Data Breakdown using 'Account Code with Descriptor' column
#             final_df['Account Code with Descriptor'] = final_df['Account Code Per Business'] + ' - ' + final_df['Account Code Per Business Description']

#             # Load the Excel template
#             template_path = '/Users/nyalpatel/Desktop/XeroAutomationWebApp/Recharging_Report_Template.xlsx'  # Replace this with the actual path to your template
#             workbook = load_workbook(template_path)
            
#             # Select the "Full Data Breakdown" sheet
#             sheet = workbook['Full Data Breakdown']  # Ensure the sheet name matches exactly

#             # Write final_df data to the "Full Data Breakdown" sheet, starting from cell A3
#             for row_idx, row in final_df.iterrows():
#                 for col_idx, value in enumerate(row):
#                     sheet.cell(row=row_idx + 4, column=col_idx + 1, value=value)  # Start writing at A3

#             # Save the modified workbook to a BytesIO stream
#             excel_stream = BytesIO()
#             workbook.save(excel_stream)
#             excel_stream.seek(0)  # Reset stream position




#             # 3. Create combined_invoices.zip in memory
#             combined_zip_stream = BytesIO()
#             with zipfile.ZipFile(combined_zip_stream, 'w', zipfile.ZIP_DEFLATED) as combined_zip:
#                 # Add sales_invoice CSV
#                 sales_invoice_csv_filename = f"sales_invoice_for_{selected_month_year_str_filenames}.csv"
#                 combined_zip.writestr(sales_invoice_csv_filename, sales_invoice_df.to_csv(index=False))
                
#                 # Add purchase_invoices.zip
#                 combined_zip.writestr('purchase_invoices.zip', zip_data1)
                
#                 # Add breakdown_invoices.zip
#                 combined_zip.writestr('breakdown_invoices.zip', zip_data2)

#                 # Add the Excel report to the ZIP
#                 combined_zip.writestr('store_company_report.xlsx', excel_stream.read())
            
#             combined_zip_stream.seek(0)

#             # Store the CSV content or ZIP data in the session after the macro is run
#             session['purchase_csv'] = purchase_invoice_df.to_csv(index=False)
#             session['breakdown_csv'] = final_df.to_csv(index=False)
#             session['sales_invoice_csv'] = sales_invoice_df.to_csv(index=False)

  


#             # Send the combined ZIP file
#             return send_file(
#                 combined_zip_stream,
#                 mimetype='application/zip',
#                 as_attachment=True,
#                 download_name='combined_invoices.zip'
#             )
        
#         except Exception as e:
#             add_log(f"Error occurred: {str(e)}", log_type="error")
#             return jsonify({"success": False, "message": "Processing failed. Please check the error log."}), 500


# @recharging_bp.route('/upload_to_xero', methods=['POST'])
# @user_login_required
# def upload_to_xero():
#     try:
#         # Retrieve the stored CSV content from the session
#         purchase_csv_content = session.get('purchase_csv')
#         breakdown_csv_content = session.get('breakdown_csv')

#         if not purchase_csv_content or not breakdown_csv_content:
#             raise ValueError("CSV content is missing in session")

#         # Process the purchase CSV into a DataFrame
#         purchase_csv_io = io.StringIO(purchase_csv_content)
#         purchase_df = pd.read_csv(purchase_csv_io)

#         # Check if the 'DueDate' column exists
#         if 'DueDate' in purchase_df.columns:
#             due_date = purchase_df['DueDate'].iloc[0]
#             if pd.isna(due_date):
#                 add_log(f"DueDate is missing for the first row in the 'purchase_csv'.", log_type="error")
#         else:
#             add_log(f"'DueDate' column is missing in the 'purchase_csv'.", log_type="error")

#         # Check if the 'InvoiceNumber' column exists
#         if 'InvoiceNumber' in purchase_df.columns:
#             invoice_number = purchase_df['InvoiceNumber'].iloc[0]
#             if pd.isna(invoice_number):
#                 add_log(f"InvoiceNumber is missing for the first row in the 'purchase_csv'.", log_type="error")
#         else:
#             add_log(f"'InvoiceNumber' column is missing in the 'purchase_csv'.", log_type="error")

#         # Process the purchase CSV into a DataFrame
#         breakdown_csv_io = io.StringIO(breakdown_csv_content)
#         breakdown_df = pd.read_csv(breakdown_csv_io)

#         # Step 1: Get the string content from the StringIO object
#         csv_string_content = breakdown_csv_io.getvalue()
#         csv_byte_content = csv_string_content.encode('utf-8')

#         invoices_data = {}

#         # Query all valid companies from TrackingCategoryModel
#         valid_companies = db.session.query(TrackingCategoryModel.tenant_name).distinct().all()
#         valid_companies = {company[0] for company in valid_companies}  # Set of valid company names


#         # Iterate over the DataFrame rows to create line items for Xero
#         for _, row in purchase_df.iterrows():
#             company_name = row.get('Company Name')

#             # Only process rows where the company name exists in TrackingCategoryModel
#             if company_name not in valid_companies:
#                 add_log(f"Company '{company_name}' not found in tracking categories. Skipping row.", log_type="error")
#                 continue

#             description = row.get('Description')
#             quantity = float(row.get('Quantity', 1))  # Default to 1 if not provided
#             unit_amount = float(row.get('UnitAmount'))
#             account_code = row.get('AccountCode')
#             tracking_option_name = row.get('TrackingOption1')  # Assuming the tracking option is in this column

#             # Query the TrackingCategoryModel to get the tracking_category_id and tracking_option_id
#             tracking_record = TrackingCategoryModel.query.filter_by(tracking_category_option=tracking_option_name).first()
#             if not tracking_record:
#                 print(tracking_option_name)
#                 add_log(f"Tracking option '{tracking_option_name}' not found for company '{company_name}'. Skipping row.", log_type="error")
#                 continue

#             # Prepare line items for each company
#             if company_name not in invoices_data:
#                 invoices_data[company_name] = []


#             # Create line item tracking
#             line_item_tracking = LineItemTracking(
#                 tracking_category_id=tracking_record.tracking_category_id,
#                 tracking_option_id=tracking_record.tracking_option_id
#             )


#             invoices_data[company_name].append(
#                 LineItem(
#                     description=description,
#                     quantity=quantity,
#                     unit_amount=unit_amount,
#                     account_code=account_code,
#                     tracking=[line_item_tracking]
#                 )
#             )

#         user = current_user
#         username = user.username
#         contact_data = get_all_contacts(user)


#         # For each company, find the tenant and contact details
#         for company, line_items in invoices_data.items():
#             file_name = (company + "Breakdown.csv")
#             company = company.strip().lower()  # Ensure company name is stripped of extra spaces and in lowercase
            
            
#             tenant_contact_info = next((tenant for tenant in contact_data if tenant['tenant_name'].lower().strip() == company), None)

#             if tenant_contact_info:
#                 tenant_id = tenant_contact_info['tenant_id']
#                 contact_record = next((contact for contact in tenant_contact_info['contacts'] if username in contact['contact_name']), None)

#                 if contact_record:
#                     contact_id = contact_record['contact_id']
#                 else:
#                     add_log(f"Contact with name {username} not found for tenant '{company}'. Skipping invoice.", log_type="error")
#                     continue
#             else:
#                 add_log(f"Tenant '{company}' not found in contact data. Skipping invoice.", log_type="error")
#                 continue

#             try:
#                 post_recharge_purchase_invoice_xero(
#                     tenant_id=tenant_id,
#                     line_items=line_items,
#                     contact_id=contact_id,  # Assuming the first contact is the main contact
#                     file_name=file_name,
#                     file_content=csv_byte_content,
#                     user = user,
#                     end_date = due_date,
#                     invoice_number = invoice_number
#                 )
            
#                 add_log(f"Invoice for company '{company}' created successfully.", log_type="success")
#             except Exception as e:
#                 add_log(f"Error creating invoice for company '{company}': {str(e)}", log_type="error")


#         return jsonify({"status": "success", "message": "Invoices uploaded to Xero successfully."})

#     except Exception as e:
#         print(f"Error in upload_to_xero: {e}")
#         add_log(f"Error in upload_to_xero: {str(e)}", log_type="errors")
#         return jsonify({"status": "error", "message": str(e)}), 500


@recharging_bp.route('/upload_to_xero', methods=['POST'])
@user_login_required
def upload_to_xero():
    user_id = current_user.id
    purchase_csv_content = session.get('purchase_csv')
    breakdown_csv_content = session.get('breakdown_csv')
    sales_invoices_csv_content = session.get('sales_invoice_csv')



    # Check if any of the required data is missing in the session
    if not purchase_csv_content or not breakdown_csv_content or not sales_invoices_csv_content:
        return jsonify({
            "status": "error",
            "message": "No files stored in the session. Please process data first before uploading to Xero."
        }), 400

    # If all required data exists, start the Celery task
    task = upload_recharge_invoices_xero_task.apply_async(args=[user_id, purchase_csv_content, breakdown_csv_content, sales_invoices_csv_content])

    return jsonify({
        "status": "success",
        "task_id": task.id,
        "message": "Upload to Xero task started."
    })

@recharging_bp.route('/download_log_csv/<task_id>', methods=['GET'])
@user_login_required
def download_log_csv(task_id):
    # Retrieve task result
    task_result = upload_recharge_invoices_xero_task.AsyncResult(task_id)

    # Ensure task completed successfully
    if task_result.state != 'SUCCESS':
        return jsonify({"status": "error", "message": "Task is not yet complete."}), 400

    # Get the successful, failed, and already processed companies from the task result
    result_data = task_result.result
    successful_companies = result_data.get('successful_companies', [])
    failed_companies = result_data.get('failed_companies', [])
    already_processed_companies = result_data.get('already_processed_companies', [])
    
    # Get the sales-related data
    sales_successful_companies = result_data.get('sales_successful_companies', [])
    sales_failed_companies = result_data.get('sales_failed_companies', [])
    sales_already_processed_companies = result_data.get('sales_already_processed_companies', [])

    # Create CSV data
    csv_output = StringIO()
    writer = csv.writer(csv_output)

    # Add a header row with metadata
    writer.writerow(['Invoice Processing Log'])
    writer.writerow([f'Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([f'Task ID: {task_id}'])
    writer.writerow([])  # Empty row for spacing

    # Add column headers for the companies
    writer.writerow(['Company Name', 'Status'])

    # Add successful companies
    if successful_companies:
        writer.writerow([])  # Empty row for spacing
        writer.writerow(['Successful Companies (Purchase Invoices)'])  # Sub-header
        for company in successful_companies:
            writer.writerow([company, 'Successful'])

    # Add failed companies
    if failed_companies:
        writer.writerow([])  # Empty row for spacing
        writer.writerow(['Failed Companies (Purchase Invoices)'])  # Sub-header
        for company in failed_companies:
            writer.writerow([company, 'Failed'])

    # Add already processed companies
    if already_processed_companies:
        writer.writerow([])  # Empty row for spacing
        writer.writerow(['Already Processed Companies (Purchase Invoices)'])  # Sub-header
        for company in already_processed_companies:
            writer.writerow([company, 'Already Processed'])

    # Add sales successful companies
    if sales_successful_companies:
        writer.writerow([])  # Empty row for spacing
        writer.writerow(['Successful Companies (Sales Invoices)'])  # Sub-header
        for company in sales_successful_companies:
            writer.writerow([company, 'Successful'])

    # Add sales failed companies
    if sales_failed_companies:
        writer.writerow([])  # Empty row for spacing
        writer.writerow(['Failed Companies (Sales Invoices)'])  # Sub-header
        for company in sales_failed_companies:
            writer.writerow([company, 'Failed'])

    # Add sales already processed companies
    if sales_already_processed_companies:
        writer.writerow([])  # Empty row for spacing
        writer.writerow(['Already Processed Companies (Sales Invoices)'])  # Sub-header
        for company in sales_already_processed_companies:
            writer.writerow([company, 'Already Processed'])

    csv_output.seek(0)  # Move the cursor to the beginning of the file

    # Create a response with the CSV data
    return Response(
        csv_output,
        mimetype='text/csv',
        headers={"Content-Disposition": "attachment;filename=invoice_processing_log.csv"}
    )




# @recharging_bp.route('/download/<filename>', methods=['GET'])
# @user_login_required
# def download_file(filename):
#     try:
#         return send_file(filename, mimetype='application/zip', as_attachment=True, download_name=filename)
#     except Exception as e:
#         return jsonify({"success": False, "message": f"Error while downloading file: {str(e)}"}), 500
    



# Route to handle file upload for recharging
@recharging_bp.route('/upload', methods=['POST'])
@user_login_required
def upload_file():
    if 'file' not in request.files:
        flash('No file part')
        add_log('File upload failed: No file part found.', log_type='error')  # Log error
        return redirect(url_for('run_macros'))

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        add_log('File upload failed: No file selected.', log_type='error')  # Log error
        return redirect(url_for('run_macros'))

    if file and file.filename.endswith(('.xls', '.xlsx', '.csv')):
        try:
            add_log(f"File upload started: {file.filename}", log_type="general")  # Log upload start

            # Save file temporarily
            temp_path = '/tmp/temp_file.xlsx' if file.filename.endswith(('.xls', '.xlsx')) else '/tmp/temp_file.csv'
            file.save(temp_path)
            add_log(f"File saved temporarily: {temp_path}", log_type="general")  # Log temp file save

            # Convert Excel to CSV if necessary
            if file.filename.endswith(('.xls', '.xlsx')):
                subprocess.run(['xlsx2csv', temp_path, '/tmp/converted_file.csv'], check=True)
                temp_path = '/tmp/converted_file.csv'  # Update path to the new CSV
                add_log(f"File converted to CSV: {temp_path}", log_type="general")  # Log conversion

            # Load CSV and remove unnecessary rows
            df = pd.read_csv(temp_path, skiprows=4)
            df = df.iloc[:-2]  # Remove specific rows
            df = df.fillna('')  # Fill NaN values with empty strings
            add_log(f"CSV file loaded and cleaned.", log_type="general")  # Log CSV load

            # Check for Southern/Northern or DOMINOS/ALL-GDK-COSTA-GYM columns
            southern_col = None
            northern_col = None

            if 'DOMINOS' in df.columns and 'OTHER' in df.columns:
                tracking_group1_col = 'DOMINOS'
                tracking_group2_col = 'OTHER'
            elif 'DOMINOS' in df.columns and 'ALL-GDK-COSTA-GYM' in df.columns:
                tracking_group1_col = 'DOMINOS'
                tracking_group2_col = 'ALL-GDK-COSTA-GYM'

            # Clear the table before adding new data
            AccountTransaction.query.delete()
            db.session.commit()  # Commit the deletion
            add_log(f"AccountTransaction table cleared.", log_type="general")  # Log table clearing

            # Store the new data in the database
            for _, row in df.iterrows():
                new_transaction = AccountTransaction(
                    user_id=session['user_id'],  # Track the current user
                    date=datetime.strptime(row['Date'], '%d %b %Y').date() if pd.notna(row['Date']) else None,  # Parse the date inline
                    source=row['Source'],
                    contact=row['Contact'] if pd.notna(row['Contact']) else None,  # Handle potential NaN values
                    description=row['Description'],
                    reference=row['Reference'],
                    debit=row['Debit'] if pd.notna(row['Debit']) else 0.0,  # Handle NaN in numerical fields
                    credit=row['Credit'] if pd.notna(row['Credit']) else 0.0,
                    gross=row['Gross'] if pd.notna(row['Gross']) else 0.0,
                    net=row['Net'] if pd.notna(row['Net']) else 0.0,
                    vat=row['VAT'] if pd.notna(row['VAT']) else 0.0,
                    account_code=int(row['Account Code']) if pd.notna(row['Account Code']) else None,  # Ensure account_code is an integer
                    account=row['Account'],
                    # Use TrackingGroup1 and TrackingGroup2 instead of Southern/Northern
                    tracking_group1=row[tracking_group1_col] if tracking_group1_col and pd.notna(row[tracking_group1_col]) else None,
                    tracking_group2=row[tracking_group2_col] if tracking_group2_col and pd.notna(row[tracking_group2_col]) else None
                )
                db.session.add(new_transaction)

            db.session.commit()  # Commit the new transactions
            flash('File successfully uploaded and data saved to the database')
            add_log(f"File {file.filename} processed successfully and data saved to the database.", log_type="general")  # Log success
        except Exception as e:
            flash(f'Error processing file: {str(e)}')
            add_log(f"Error processing file {file.filename}: {str(e)}", log_type="error")  # Log error
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
                add_log(f"Temporary file {temp_path} removed.", log_type="general")  # Log temp file cleanup
    else:
        flash('Invalid file format')
        add_log(f"Invalid file format: {file.filename}", log_type="error")  # Log invalid format

    return redirect(url_for('recharging.recharging_home'))

