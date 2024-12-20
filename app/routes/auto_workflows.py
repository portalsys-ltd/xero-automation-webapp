# app/routes/auto_workflows.py

from flask import Blueprint, jsonify, session, request, current_app
from app.xero import xero_token_required, aggregate_auto_workflows_data, get_inbox_files_from_management_company
from app.routes.logs import add_log
from xero_python.exceptions import HTTPStatusException
from flask_login import current_user
from app.celery_tasks import pre_process_dom_purchase_invoices_task, process_dom_purchase_invoices_task, process_dom_sales_invoices_task, process_cocacola_task, process_eden_farm_task, process_textman_task, update_invoice_record_task, process_inventory_task, create_inventory_journals
from app.models import TrackingCategoryModel, User, DomPurchaseInvoicesTenant, InvoiceRecord, TaskStatus, InventoryRecord
import pandas as pd
import re
import json
from app.routes.auth import user_login_required
from app import db 
from celery.result import AsyncResult


auto_workflows_bp = Blueprint('auto_workflows', __name__, url_prefix='/auto_workflows')


@auto_workflows_bp.route('/check_pre_processed_invoices', methods=['GET'])
@user_login_required
def check_pre_processed_invoices():
    user_id = current_user.id
    # Retrieve the user object from the database
    user = User.query.get(user_id)
    data, management_tenant_id = get_inbox_files_from_management_company(user)


    # Step 1: Identify live tenants and get expected store numbers for each tenant
    live_tenants = DomPurchaseInvoicesTenant.query.filter_by(user_id=user_id).all()
    expected_store_numbers = []

    # Dictionary to hold tenant names and their expected store numbers
    tenant_store_numbers = {}

    for tenant in live_tenants:
        tenant_name = tenant.tenant_name
        store_numbers = TrackingCategoryModel.query.filter_by(user_id=user_id, tenant_name=tenant_name).all()
        store_numbers_list = [store.store_number for store in store_numbers if store.store_number]
        tenant_store_numbers[tenant_name] = store_numbers_list
        expected_store_numbers.extend(store_numbers_list)  # Collect all store numbers for the user

    # Dictionary to store files by store number and week
    grouped_files = {}

    for file in data[0]['files']:
        file_id = file['file_id']
        file_name = file['file_name']
        mime_type = file['mime_type']

        # Check if the file is pre-processed
        if not file_name.startswith("PRE PROCESSED"):
            continue

        # Extract the store number from the file name (format: S-XXXXX)
        store_number_match = re.search(r"S-(\d{5})", file_name)
        store_number = store_number_match.group(1) if store_number_match else None
        if not store_number:
            print(f"No store number found in file '{file_name}', skipping...")
            continue

        # Extract the statement date from the file name (assuming it's in the format `DD-MM-YYYY`)
        statement_date_match = re.search(r"(\d{2}-\d{2}-\d{4})", file_name)
        statement_date = statement_date_match.group(1) if statement_date_match else None
        if not statement_date:
            print(f"No statement date found in file '{file_name}', skipping...")
            continue

        # Convert statement date to week identifier (e.g., ISO week)
        week_identifier = pd.to_datetime(statement_date, format='%d-%m-%Y').isocalendar()[1]
        
        # Create a unique key based on week identifier
        key = week_identifier

        # Initialize the week key in the dictionary if not present
        if key not in grouped_files:
            grouped_files[key] = {}

        # Initialize the store number entry for this week
        if store_number not in grouped_files[key]:
            grouped_files[key][store_number] = {'csv': None, 'pdf': None}

        # Group the files by type under the store number for this week
        if mime_type == 'text/csv' or file_name.lower().endswith('.csv'):
            grouped_files[key][store_number]['csv'] = file
        elif mime_type == 'application/pdf' or file_name.lower().endswith('.pdf'):
            grouped_files[key][store_number]['pdf'] = file

    # Structure the result dictionary for each week
    weeks_summary = {}
    for week, stores in grouped_files.items():
        # Check for matched pairs
        matched_pairs = []
        missing_stores = []
        for store_number in expected_store_numbers:
            file_pair = stores.get(store_number)
            if file_pair and file_pair['csv'] and file_pair['pdf']:
                matched_pairs.append({
                    'store_number': store_number,
                    'statement_date': file_pair['csv']['file_name'].split('_')[-1],  # Extract date from filename
                    'csv_file_id': file_pair['csv']['file_id'],
                    'pdf_file_id': file_pair['pdf']['file_id'],
                    'csv_file_name': file_pair['csv']['file_name'],
                    'pdf_file_name': file_pair['pdf']['file_name']
                })
            else:
                # Track missing store numbers
                missing_stores.append(store_number)

        # Add to weekly summary
        weeks_summary[week] = {
            'matched_invoice_count': len(matched_pairs),
            'matched_pairs': matched_pairs,
            'expected_invoice_count': len(expected_store_numbers),
            'missing_store_numbers': missing_stores
        }


    return jsonify({
        'status': 'success',
        'message': 'Pre-processed purchase invoices checked and grouped successfully.',
        'weeks_summary': weeks_summary
    })


@auto_workflows_bp.route('/data', methods=['GET'])
@user_login_required
def auto_workflows_data():
    if not current_user:
        add_log("User ID not found in session.", log_type="errors")
        return jsonify({"error": "User not authenticated."}), 401
    
    try:
        data = aggregate_auto_workflows_data(current_user)
        return data
    except Exception as e:
        add_log(f"Error aggregating auto workflows data: {e}", log_type="errors")
        return jsonify({"error": "Failed to fetch auto workflows data."}), 500


@auto_workflows_bp.route('/pre_process_dom_purchase_invoices', methods=['GET'])
@user_login_required
def pre_process_dom_purchase_invoices_route():
    user_id = current_user.id  # Pass the user's ID instead of the whole object

    task_type = "pre_process_dom_purchase_invoices"
    task = pre_process_dom_purchase_invoices_task.apply_async(args=[user_id])
    # Save task details to the database
    try:
        new_task = TaskStatus(
            task_id=task.id,
            user_id=user_id,
            task_type=task_type,
            status='in_progress'
        )
        db.session.add(new_task)
        db.session.commit()

        print(f"TaskStatus entry created for task_id: {task.id}")
    except Exception as e:
        print(f"Error saving TaskStatus to database: {e}")
        return jsonify({"status": "error", "message": "Failed to save task to database."}), 500

  

    if task:
        return jsonify({'task_id': task.id, 'message': 'Task started!'})
    else:
        return jsonify({"error": "Task failed to start"}), 500  
    


@auto_workflows_bp.route('/process_dom_purchase_invoices', methods=['GET'])
@user_login_required
def process_dom_purchase_invoices_route():
    user_id = current_user.id
    week = request.args.get('week')  # Get the week parameter from the query string

    if not week:
        return jsonify({"error": "Week parameter is missing"}), 400

    task = process_dom_purchase_invoices_task.apply_async(args=[user_id, int(week)])
    
    task_type = "process_dom_purchase"

    # Save task details to the database
    try:
        new_task = TaskStatus(
            task_id=task.id,
            user_id=user_id,
            task_type=task_type,
            status='in_progress'
        )
        db.session.add(new_task)
        db.session.commit()

        print(f"TaskStatus entry created for task_id: {task.id}")
    except Exception as e:
        print(f"Error saving TaskStatus to database: {e}")
        return jsonify({"status": "error", "message": "Failed to save task to database."}), 500


    if task:
        return jsonify({'task_id': task.id, 'message': 'Task started!'})
    else:
        return jsonify({"error": "Task failed to start"}), 500

    
@auto_workflows_bp.route('/process_dom_sales_invoices', methods=['GET'])
@user_login_required
def process_dom_sales_invoices_route():
    user_id = current_user.id  # Pass the user's ID instead of the whole object
    task = process_dom_sales_invoices_task.apply_async(args=[user_id])

    task_type = "process_dom_sales"

    # Save task details to the database
    try:
        new_task = TaskStatus(
            task_id=task.id,
            user_id=user_id,
            task_type=task_type,
            status='in_progress'
        )
        db.session.add(new_task)
        db.session.commit()

        print(f"TaskStatus entry created for task_id: {task.id}")
    except Exception as e:
        print(f"Error saving TaskStatus to database: {e}")
        return jsonify({"status": "error", "message": "Failed to save task to database."}), 500


    if task:
        return jsonify({'task_id': task.id, 'message': 'Task started!'})
    else:
        return jsonify({"error": "Task failed to start"}), 500
    


@auto_workflows_bp.route('/process_coca-cola_invoices', methods=['POST'])
@user_login_required
def coca_cola_process_invoices_route():
    user_id = current_user.id  # Pass the user's ID instead of the whole object
    task = process_cocacola_task.delay(user_id)
    return jsonify({"message": "Coca-Cola invoice processing has started.", "task_id": task.id}), 202  # 202 indicates the task is accepted for processing


@auto_workflows_bp.route('/process_text-man_invoices', methods=['POST'])
@user_login_required
def text_man_process_invoices_route():
    user_id = current_user.id  # Pass the user's ID instead of the whole object
    task = process_textman_task.delay(user_id)
    return jsonify({"message": "Text Management invoice processing has started.", "task_id": task.id}), 202


@auto_workflows_bp.route('/process_eden-farm_invoices', methods=['POST'])
@user_login_required
def eden_farm_process_invoices_route():
    user_id = current_user.id  # Pass the user's ID instead of the whole object
    task = process_eden_farm_task.delay(user_id)
    return jsonify({"message": "Eden Farm invoice processing has started.", "task_id": task.id}), 202
    

@auto_workflows_bp.route('/active_task_status', methods=['GET'])
@user_login_required
def active_task_status():
    user_id = current_user.id

    # Query the most recent tasks for the user, restricted to specific task types
    tasks = (
        TaskStatus.query.filter(
            TaskStatus.user_id == user_id,
            TaskStatus.task_type.in_(['pre_process_dom_purchase_invoices', 'process_dom_purchase', 'sales', 'process_dom_sales','process_inventory']),  # Filter by task types
            TaskStatus.status.in_(['in_progress', 'failed', 'completed'])  # Include relevant statuses
        )
        .order_by(TaskStatus.created_at.desc())  # Order by most recent
        .all()
    )

    if not tasks:
        return jsonify({'status': 'no_active_task'})

    # Initialize variables to track the most recent tasks by status
    most_recent_completed = None
    most_recent_failed_or_in_progress = None

    # Iterate through tasks to find the most recent ones for relevant statuses
    for task in tasks:
        if task.status == 'completed' and not most_recent_completed:
            most_recent_completed = task
        elif task.status in ['in_progress', 'failed'] and not most_recent_failed_or_in_progress:
            most_recent_failed_or_in_progress = task

        # If we have both conditions satisfied, we can stop checking further
        if most_recent_completed and most_recent_failed_or_in_progress:
            break

    # If there's a completed task that is more recent than any failed or in-progress task, return no active task
    if (most_recent_completed and most_recent_failed_or_in_progress and 
            most_recent_completed.created_at > most_recent_failed_or_in_progress.created_at):
        return jsonify({'status': 'no_active_task'})

    # Otherwise, return the most recent failed or in-progress task
    if most_recent_failed_or_in_progress:
        return jsonify({
            'task_id': most_recent_failed_or_in_progress.task_id,
            'task_type': most_recent_failed_or_in_progress.task_type,
            'status': most_recent_failed_or_in_progress.status,
            'progress': 0 if most_recent_failed_or_in_progress.status == 'failed' else most_recent_failed_or_in_progress.progress,
            'message': 'Task failed.' if most_recent_failed_or_in_progress.status == 'failed' else 'Task is in progress...',
            'state': 'FAILURE' if most_recent_failed_or_in_progress.status == 'failed' else 'PROGRESS',
        })

    # If no relevant tasks are found, return no active task
    return jsonify({'status': 'no_active_task'})


@auto_workflows_bp.route('/task_status/<task_id>', methods=['GET'])
def task_status(task_id):

    task = AsyncResult(task_id) 

    print(task.id)

     # Default response
    response = {
        'state': task.state,
        'progress': 0,
        'message': 'Task pending or unknown.',
        'errors': []
    }
 
    if task.state == 'PENDING':
        response = {
            'state': task.state,
            'progress': 0
        }
    elif task.state == 'PROGRESS':
        response = {
            'state': task.state,
            'current': task.info.get('current', 0),
            'total': task.info.get('total', 1),
            'progress': int((task.info.get('current', 0) / task.info.get('total', 1)) * 100)
        }
    elif task.state == 'SUCCESS':
        # Handle the case where task.result or task.info is None
        result = task.result if isinstance(task.result, dict) else {}
        response = {
            'state': task.state,
            'progress': 100,  # Task is complete, so progress is 100%
            'message': result.get('message', 'Task completed successfully'),
            'errors': result.get('errors', [])  # Ensure errors are included in the response
        }
    elif task.state == 'FAILURE':
        response = {
            'state': task.state,
            'progress': 0,
            'errors': str(task.info),  # The error message from the task
            'traceback': str(task.traceback)  # You can include the traceback for debugging purposes
        }
    else:
        response = {
            'state': task.state,
            'progress': 0,
            'message': 'Unknown task state',
            'errors': []
        }

    return jsonify(response)



@auto_workflows_bp.route('/start_update_invoice_record', methods=['POST'])
@user_login_required
def start_update_invoice_record():
    try:
        print("Route triggered successfully!")  # Debugging statement
        
        # Retrieve the current user
        user_id = current_user.id
        print(f"Current User ID: {user_id}")  # Debugging

        if not user_id:
            return jsonify({"error": "User not found"}), 404

        # Trigger the Celery task
        task = update_invoice_record_task.apply_async(args=[user_id])
        print(f"Task {task.id} started for user {user_id}")  # Debugging task trigger

        return jsonify({"task_id": task.id, "message": "Task triggered successfully"}), 202

    except Exception as e:
        print(f"Error occurred: {str(e)}")  # Print error to console
        return jsonify({"error": str(e)}), 500





@auto_workflows_bp.route('/load_invoice_summary', methods=['POST'])
@user_login_required
def load_invoice_summary():
    try:
        # Get week and year from the request body
        data = request.get_json()
        week_number = data.get("week_number")
        year = data.get("year")


        # Validate inputs
        if not week_number or not year:
            return jsonify({"error": "Week and year are required"}), 400

        # Fetch all tenant names for the current user
        tenant_names = [
            tenant.tenant_name
            for tenant in DomPurchaseInvoicesTenant.query.filter_by(user_id=current_user.id).all()
        ]

    
        # Fetch all stores for those tenants from TrackingCategoryModel
        store_details = TrackingCategoryModel.query.filter(
            TrackingCategoryModel.tenant_name.in_(tenant_names)
        ).all()

        print(store_details)

        # Prepare store data for matching
        store_records = []
        for store in store_details:

            # Check for a matching record in InvoiceRecord
            existing_record = InvoiceRecord.query.filter_by(
                week_number=week_number,
                year=year,
                store_number=store.store_number
            ).first()

            # Determine status: green tick or red X
            status = "✅" if existing_record else "❌"

            if not store.store_number or not re.match(r"^\d{5}$", store.store_number):
                continue
                
            # Append store record to the response list
            store_records.append({
                "store_name": store.tracking_category_option,  # Corrected column
                "store_number": store.store_number,
                "tenant_name": store.tenant_name,  # Use this as a placeholder if needed
                "status": status
            })

        return jsonify({"data": store_records}), 200

    except Exception as e:
        print(f"Error loading invoice summary: {str(e)}")
        return jsonify({"error": "Failed to load summary"}), 500
    

@auto_workflows_bp.route('/load_invoice_summary_sales', methods=['POST'])
@user_login_required
def load_invoice_summary_sales():
    try:
        import re

        # Get week and year from the request body
        data = request.get_json()
        week_number = data.get("week_number")
        year = data.get("year")

        # Validate inputs
        if not week_number or not year:
            return jsonify({"error": "Week and year are required"}), 400

        # Fetch all tenant names for the current user
        tenant_names = [
            tenant.tenant_name
            for tenant in DomPurchaseInvoicesTenant.query.filter_by(user_id=current_user.id).all()
        ]

        # Fetch all stores for those tenants from TrackingCategoryModel
        store_details = TrackingCategoryModel.query.filter(
            TrackingCategoryModel.tenant_name.in_(tenant_names)
        ).all()

        # Prepare store data for matching
        store_records = []
        for store in store_details:
            # Ignore invalid store numbers
            if not store.store_number or not re.match(r"^\d{5}$", store.store_number):
                continue

            # Check for mileage record
            mileage_record = InvoiceRecord.query.filter_by(
                week_number=week_number,
                year=year,
                store_number=store.store_number,
                invoice_type="mileage"
            ).first()

            # Check for sales record
            sales_record = InvoiceRecord.query.filter_by(
                week_number=week_number,
                year=year,
                store_number=store.store_number,
                invoice_type="sales"
            ).first()

            # Append store record to the response list
            store_records.append({
                "store_name": store.tracking_category_option,
                "store_number": store.store_number,
                "tenant_name": store.tenant_name,
                "mileage_status": "✅" if mileage_record else "❌",
                "sales_status": "✅" if sales_record else "❌"
            })

        return jsonify({"data": store_records}), 200

    except Exception as e:
        print(f"Error loading invoice summary: {str(e)}")
        return jsonify({"error": "Failed to load summary"}), 500



@auto_workflows_bp.route('/process_inventory_data', methods=['GET'])
@user_login_required
def process_inventory_data_route():
    user_id = current_user.id  # Get the current user ID

    # Trigger the inventory processing Celery task
    task = process_inventory_task.apply_async(args=[user_id])

    # Save task details in the database for tracking
    try:
        new_task = TaskStatus(
            task_id=task.id,
            user_id=user_id,
            task_type="process_inventory",
            status='in_progress'
        )
        db.session.add(new_task)
        db.session.commit()
        print(f"TaskStatus entry created for inventory task_id: {task.id}")
    
    except Exception as e:
        print(f"Error saving TaskStatus to database: {e}")
        return jsonify({"status": "error", "message": "Failed to save task to database."}), 500

    return jsonify({"task_id": task.id, "message": "Inventory processing task started!"}), 202


import os
from flask import send_file, session
from app.models import TrackingCategoryModel
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter


@auto_workflows_bp.route('/create_inventory_template', methods=['GET'])
@user_login_required
def create_inventory_template():
    user_id = current_user.id

    # Step 1: Query store names for the current user and filter later for digit-only store numbers
    stores = TrackingCategoryModel.query.filter_by(user_id=user_id).filter(
        TrackingCategoryModel.store_contact.isnot(None)
    ).all()

    # Filter for store numbers containing only digits
    stores = [store for store in stores if store.store_number and store.store_number.isdigit()]

    # Step 2: Create an Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Template"

    # Add headers
    ws['A1'] = "Last Day of Month"
    ws['B1'] = ""
    ws['A2'] = datetime.now().strftime("%d/%m/%Y")  # Default today's date for reference

    ws['A4'] = "Store List"
    ws['B4'] = "Amount"

    # Add store data
    row = 5
    for store in stores:
        ws.cell(row=row, column=1, value=store.tracking_category_option)  # Store Name
        ws.cell(row=row, column=2, value="")  # Leave Amount blank
        row += 1

    # Set the width of the first column
   
    ws.column_dimensions[get_column_letter(1)].width = 25  # Adjust width as needed


    # Step 3: Lock the "Store List" column and protect the worksheet
    for row in ws.iter_rows(min_row=5, max_row=4 + len(stores), min_col=1, max_col=1):
        for cell in row:
            cell.protection = Protection(locked=True)  # Lock the cell

    # Unlock the "Amount" column so it can be edited
    ws['A2'].protection = Protection(locked=False)  # Unlock the date cell
    for row in ws.iter_rows(min_row=5, max_row=4 + len(stores), min_col=2, max_col=2):
        for cell in row:
            cell.protection = Protection(locked=False)  # Unlock the cell

    # Protect the worksheet
    ws.protection.enable()  # Enable worksheet protection
    ws.protection.password = "stock"  # Set a password for the protection

    # Step 4: Save the file temporarily in the temp_files directory
    temp_dir = os.path.join(current_app.root_path, "static", "temp_files")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    file_name = f"Stock_Record.xlsx"
    file_path = os.path.join(temp_dir, file_name)

    wb.save(file_path)

    # Step 5: Send file for download
    return send_file(file_path, as_attachment=True, download_name=file_name)


@auto_workflows_bp.route('/get_inventory_status', methods=['GET'])
def get_inventory_status():
    user_id = session.get('user_id')  # Assuming user_id is stored in the session

    if not user_id:
        return jsonify({"error": "User is not logged in"}), 401

    records = InventoryRecord.query.with_entities(
        InventoryRecord.month,
        InventoryRecord.year,
        db.func.count(InventoryRecord.id).label('record_count'),
        db.func.sum(InventoryRecord.processed.cast(db.Integer)).label('processed_count')
    ).filter_by(user_id=user_id).group_by(InventoryRecord.month, InventoryRecord.year).all()

    result = [
        {
            "month": record.month,
            "year": record.year,
            "data_complete": record.record_count > 0,  # Example logic for completeness
            "processed": record.processed_count == record.record_count,  # All processed
            "processed_count": record.processed_count or 0,  # Handle null case
            "total_count": record.record_count
        }
        for record in records
    ]
    return jsonify(result)


@auto_workflows_bp.route('/get_unprocessed_stores', methods=['GET'])
def get_unprocessed_stores():
    user_id = session.get('user_id')  # Assuming user_id is stored in the session

    if not user_id:
        return jsonify({"error": "User is not logged in"}), 401

    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)

    if not month or not year:
        return jsonify({"error": "Month and year are required"}), 400

    unprocessed_stores = InventoryRecord.query.filter_by(
        user_id=user_id,
        month=month,
        year=year,
        processed=False
    ).with_entities(InventoryRecord.store_name).distinct().all()

    return jsonify({
        "month": month,
        "year": year,
        "unprocessed_stores": [store.store_name for store in unprocessed_stores]
    })




from datetime import datetime, timedelta
from flask import jsonify, request



@auto_workflows_bp.route('/process_inventory_month', methods=['POST'])
@user_login_required
def process_inventory_month():

    user_id = current_user.id
    data = request.json
    month = data.get('month')
    year = data.get('year')


    if not month or not year or not user_id:
        return jsonify({"error": "Month, year, and user ID are required."}), 400

    # Convert to datetime for easier manipulation
    current_month_date = datetime(year, month, 1)
    previous_month_date = current_month_date - timedelta(days=1)

    # Get the previous month and year
    previous_month = previous_month_date.month
    previous_year = previous_month_date.year

    # Query inventory records for the current and previous month
    current_month_records = InventoryRecord.query.filter_by(
        month=month, year=year, user_id=user_id
    ).all()

    previous_month_records = InventoryRecord.query.filter_by(
        month=previous_month, year=previous_year, user_id=user_id
    ).all()

    if not current_month_records:
        return jsonify({"error": "No records found for the specified month and year."}), 404

    if not previous_month_records:
        return jsonify({"error": "No records found for the previous month."}), 404

    task = create_inventory_journals.apply_async(
        args=[serialize_records(current_month_records), serialize_records(previous_month_records), user_id, month, year]
    )


    return jsonify({"success": True, "task_id": task.id})


def serialize_records(records):
    """Helper function to serialize inventory records for the Celery task."""
    return [
        {
            "id": record.id,
            "store_name": record.store_name,
            "amount": record.amount,
            "processed": record.processed,
            "month": record.month,
            "year": record.year,
        }
        for record in records
    ]
