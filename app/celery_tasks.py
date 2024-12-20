# app/celert_tasks.py
import os
from app import celery
from app.routes.logs import add_log
from flask import app
import time
from app.models import *
from xero_python.accounting import AccountingApi, ContactPerson, Contact, Contacts, Invoices, Invoice, LineItem, Contact, LineItemTracking
from collections import defaultdict
from sqlalchemy import and_
from decimal import Decimal, ROUND_HALF_UP
import re
from collections import defaultdict
import pandas as pd
import pdfplumber
import csv
from celery import shared_task
from io import StringIO, BytesIO
from sqlalchemy import func, and_



from flask import session, jsonify, current_app
from app.models import TrackingCategoryModel, User
import io
import base64

from app.xero import (
    get_tracking_categories_from_xero,
    fetch_dom_invoicing_data,
    fetch_file_content,
    bulk_create_bills,
    move_and_rename_file,
    create_folder_if_not_exists,
    extract_statement_date_from_pdf,
    extract_statement_date_from_csv,
    fetch_dom_management_company_data,
    post_dom_sales_invoice_with_attachment,
    get_all_contacts,
    get_invoices_and_credit_notes,
    extract_eden_farm_invoice_data,
    extract_coca_cola_invoice_data,
    extract_textman_invoice_data,
    convert_invoice_to_credit_memo,
    convert_credit_memo_to_invoice,
    assign_tracking_code_to_invoice,
    assign_tracking_code_to_credit_note,
    post_recharge_purchase_invoice_xero,
    post_recharge_sales_invoice_xero, 
    get_inbox_files_from_management_company,
    rename_file,
    update_invoice_records,
    process_inventory_file,
    create_inventory_journals_in_xero,
    refresh_xero_token,
    void_invoice

)


@shared_task(bind=True, name='app.celery_tasks.pre_process_dom_purchase_invoices_task')
def pre_process_dom_purchase_invoices_task(self, user_id):
    user = User.query.get(user_id)

    # Fetch the TaskStatus record for the current task
    task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

    if task_status:
        task_status.status = 'in_progress'
        task_status.result = "task_started"
        db.session.commit()

    data, management_tenant_id = get_inbox_files_from_management_company(user)

    
    # Initialize an error list to collect issues with nominal codes
    error_messages = []

    processed_file_names = set()  # Track processed file names

    if data:
        tenant_id = data[0]['tenant_id']

        processed_folder = create_folder_if_not_exists('Dominos Supplier Invoices - Processed', tenant_id, user)
        rejected_folder = create_folder_if_not_exists('Dominos Supplier Invoices - Rejected', tenant_id, user)

        rejected_folder_id = rejected_folder.id
        processed_folder_id = processed_folder.id

   

        for file in data[0]['files']:
            file_id = file['file_id']
            file_name = file['file_name']
            mime_type = file['mime_type']
            statement_date = None
            all_codes_exist = True

            # Check if the file name contains 'CustAccountStatementExt.Report'
            if 'CustAccountStatementExt.Report' not in file_name:
                #print(f"Skipping file '{file_name}' as it does not contain 'CustAccountStatementExt.Report'")
                continue

            # Skip files already marked as pre-processed
            if file_name.startswith("PRE PROCESSED"):
                #print(f"Skipping already pre-processed file: {file_name}")
                continue

            # Adjust MIME type based on file extension if incorrect
            if mime_type == 'application/octet-stream':
                if file_name.lower().endswith('.csv'):
                    mime_type = 'text/csv'
                elif file_name.lower().endswith('.pdf'):
                    mime_type = 'application/pdf'
            
            # Extract store number in format "S-XXXXX" from file name
            store_number_match = re.search(r"S-(\d{5})", file_name)
            
            if store_number_match:
                store_number = store_number_match.group(1)
                store_record = TrackingCategoryModel.query.filter_by(user_id=user_id, store_number=store_number).first()

                if not store_record:
                    error_messages.append(f"Store number '{store_number}' in file '{file_name}' does not exist in the database.")
                else:
                    # Check if tenant_name from TrackingCategoryModel exists in DomPurchaseInvoicesTenant
                    tenant_name = store_record.tenant_name
                    tenant_exists = DomPurchaseInvoicesTenant.query.filter_by(user_id=user_id, tenant_name=tenant_name).first()

                    if not tenant_exists:
                        error_messages.append(f"Tenant name '{tenant_name}' for store number '{store_number}' has not been set live.")
            else:
                error_messages.append(f"No valid store number found in file '{file_name}'.")
                continue


            # Process PDF files
            if mime_type == 'application/pdf' or file_name.lower().endswith('.pdf'):
                statement_date = extract_statement_date_from_pdf(tenant_id, file_id, user)
                


       
            # Process CSV files
            elif mime_type == 'text/csv' or file_name.lower().endswith('.csv'):
                statement_date = extract_statement_date_from_csv(tenant_id, file_id, user)
                csv_content = fetch_file_content(user, tenant_id, file_id)
                df = pd.read_csv(csv_content)

                # Check if 'Nominal code' column exists
                if 'Nominal code' not in df.columns:
                    error_messages.append(f"Missing 'Nominal code' column in file '{file_name}'.")
                    all_codes_exist = False
                    continue

                # Check if 'Store' column exists
                if 'Store' not in df.columns:
                    error_messages.append(f"Missing 'Store' column in file '{file_name}'.")
                    all_codes_exist = False
                    continue

                # Loop through each row in the CSV data
                for index, row in df.iterrows():
                    # Check for nominal code
                    nominal_code = row.get('Nominal code')  # Use get() to handle missing values
                    if pd.isna(nominal_code):
                        error_messages.append(f"Missing nominal code in file '{file_name}' at row {index + 1}.")
                        all_codes_exist = False
                        continue

                    # Check if the nominal code exists in the database
                    code_exists = DomNominalCodes.query.filter_by(user_id=user.id, nominal_code=str(nominal_code)).first()
                    if not code_exists:
                        error_messages.append(f"Dominos Group Nominal code '{nominal_code}' does not exist in the database, please add to DOM invoice settings.")
                        all_codes_exist = False
                        continue

                    # Check for store code
                    store_code = str(row.get('Store')).strip()
                    if pd.isna(store_code):
                        error_messages.append(f"Missing store code in file '{file_name}' at row {index + 1}.")
                        all_codes_exist = False
                        continue

                    # Verify store code exists in the TrackingCategoryModel
                    store_exists = TrackingCategoryModel.query.filter_by(user_id=user.id, store_number=store_code).first()
                    if not store_exists:
                        all_codes_exist = False
                        error_messages.append(f"Store code '{store_code}' in file '{file_name}' at row {index + 1} does not exist in the database. Please add this store code to the DOM invoice settings.")


            not_duplicate = True

            # If store_code and statement_date exist, check for duplicates
            if store_number and statement_date:
                # Base file name without extensions
                base_file_name = file_name.replace('.csv', '').replace('.pdf', '')

                processed_log = LogEntry.query.filter(
                    and_(
                        LogEntry.message.like(f"%Store code '{store_number}'%"),
                        LogEntry.message.like(f"%Statement Date '{statement_date}'%"),
                        LogEntry.log_type == "dom_purchase_invoice",
                        LogEntry.user_id == user_id
                    )
                ).first()

                # If already processed or duplicate, move to rejected
                if processed_log or file_name in processed_file_names:

                    add_log(f"File {file_name} (Store: {store_number}, Statement Date: {statement_date}) has already been processed or is a duplicate.",
                            log_type="general", user_id=user_id)
                    rejected_filename = f"{os.path.splitext(file_name)[0]} (Rejected as duplicate or already processed){os.path.splitext(file_name)[1]}"
                    
                    move_and_rename_file(file_id, rejected_folder_id, rejected_filename, mime_type, user, tenant_id)

                    error_messages.append(f"{file_name} has already been processed or is a duplicate, file has been moved to rejected with new file name: {rejected_filename}")
                    
                    not_duplicate = False

        
                # Track the processed file to prevent duplicates in the same batch
                processed_file_names.add(file_name)


            # Rename file to include "Pre Processed" and statement date if it was extracted
            if all_codes_exist and statement_date and store_record and tenant_exists and not_duplicate:
                safe_statement_date = statement_date.replace('/', '-')
                base_name, extension = file_name.rsplit('.', 1)
                new_file_name = f"PRE PROCESSED - {base_name}_{safe_statement_date}.{extension}"

                rename_successful = rename_file(file_id, new_file_name, file_type=mime_type, user=user, tenant_id=tenant_id)
                if rename_successful:
                    print(f"File renamed to {new_file_name}")
                else:
                    error_messages.append(f"Failed to rename file {file_name} to {new_file_name}")


    if error_messages:
        task_status.status = 'failed'
        task_status.result = "\n".join(error_messages)
        
    else:
        task_status.status = 'completed'
        task_status.result = "Task completed successfully."
    
    db.session.commit()
            

    return {
        'status': 'success' if not error_messages else 'error',
        'message': 'Processing complete.' if not error_messages else 'Processing completed with some errors.',
        'errors': error_messages
    }

@shared_task(bind=True, name='app.celery_tasks.process_dom_purchase_invoices_task')
def process_dom_purchase_invoices_task(self, user_id, week):

    # Fetch the TaskStatus record for the current task
    task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

    if task_status:
        task_status.status = 'in_progress'
        task_status.result = "task_started"
        db.session.commit()  
    

    task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

    # Load tracking codes from the database
    tracking_codes = {}
    categories = TrackingCategoryModel.query.filter_by(user_id=user_id).all()

    for category in categories:
        if category.store_number:  # Only add entries with a valid store number
            tracking_codes[category.store_number] = {
                "tracking_category_id": category.tracking_category_id,
                "tracking_option_id": category.tracking_option_id
            }
            
    add_log(f"Tracking codes loaded from the database", log_type="general", user_id=user_id)


    # Retrieve the user object from the database
    user = User.query.get(user_id)
    data, management_tenant_id = get_inbox_files_from_management_company(user)
    filtered_list = []
    processed_file_count = 0

    # Step 1: Identify live tenants and get expected store numbers for each tenant
    live_tenants = DomPurchaseInvoicesTenant.query.filter_by(user_id=user_id).all()
    expected_stores = []

    # Dictionary to hold tenant and their expected store numbers
    tenant_store_numbers = {}
    tenant_names_map = {}

    for tenant in live_tenants:
        tenant_name = tenant.tenant_name
        store_numbers = TrackingCategoryModel.query.filter_by(user_id=user_id, tenant_name=tenant_name).all()
        store_numbers_list = [store.store_number for store in store_numbers if store.store_number]
        tenant_store_numbers[tenant_name] = store_numbers_list
        expected_stores.extend(store_numbers_list)  # Add all stores for the user

        # Map store numbers to tenant name
        for store_number in store_numbers_list:
            tenant_names_map[store_number] = tenant_name

    # Fetch tenant IDs for easy lookup
    tenant_id_map = {tenant.tenant_name: tenant.tenant_id for tenant in XeroTenant.query.filter_by(user_id=user_id).all()}

    # Dictionary to store files by store number for the specified week
    grouped_files = {}

    for file in data[0]['files']:
        file_id = file['file_id']
        file_name = file['file_name']
        mime_type = file['mime_type']

        # Check if the file is pre-processed
        if not file_name.startswith("PRE PROCESSED"):
            continue

        # Extract the store number from the file name (format: S-XXXXX)
        store_number_match = re.search(r"S-(\d{5})", file_name)
        store_number = store_number_match.group(1) if store_number_match else None
        if not store_number:
            print(f"No store number found in file '{file_name}', skipping...")
            continue

        # Extract the statement date from the file name (assuming it's in the format `DD-MM-YYYY`)
        statement_date_match = re.search(r"(\d{2}-\d{2}-\d{4})", file_name)
        statement_date = statement_date_match.group(1) if statement_date_match else None
        if not statement_date:
            print(f"No statement date found in file '{file_name}', skipping...")
            continue

        # Convert statement date to week identifier (e.g., ISO week)
        week_identifier = pd.to_datetime(statement_date, format='%d-%m-%Y').isocalendar()[1]
        
        # Only process files for the specified week
        if week_identifier != week:
            continue

        # Initialize the store number entry for this week in grouped_files
        if store_number not in grouped_files:
            grouped_files[store_number] = {'csv': None, 'pdf': None}

        # Group the files by type under the store number for this week
        if mime_type == 'text/csv' or file_name.lower().endswith('.csv'):
            grouped_files[store_number]['csv'] = file
        elif mime_type == 'application/pdf' or file_name.lower().endswith('.pdf'):
            grouped_files[store_number]['pdf'] = file

    # Process files for each store in the specified week
    matched_files = []
    for store_number, file_pair in grouped_files.items():
        if file_pair['csv'] and file_pair['pdf']:
            tenant_name = tenant_names_map.get(store_number)
            tenant_id = tenant_id_map.get(tenant_name)

            matched_files.append({
                'tenant_id': tenant_id,
                'tenant_name': tenant_name,
                'store_number': store_number,
                'week': week,
                'statement_date': file_pair['csv']['file_name'].split('_')[-1],  # Extract date from filename
                'csv_file': file_pair['csv'],
                'pdf_file': file_pair['pdf']
            })

    # Populate the filtered list with matched pairs
    for match in matched_files:
        filtered_list.append({
            'tenant_id': match['tenant_id'],
            'tenant_name': match['tenant_name'],
            'store_code': match['store_number'],
            'week': match['week'],
            'statement_date': match['statement_date'],
            'csv_file_id': match['csv_file']['file_id'],
            'pdf_file_id': match['pdf_file']['file_id'],
            'csv_file_name': match['csv_file']['file_name'],
            'pdf_file_name': match['pdf_file']['file_name'],
        })

    total_files = len(filtered_list)

    # Create a list to store the processed data
    processed_files = []

    # Initialize tracking variables
    current_month = None

    current_reference = "A"  # Start with "A"


    # Iterate over the filtered files
    for file_data in filtered_list:

        tenant_id = file_data['tenant_id']
        tenant_name = file_data['tenant_name']
        csv_file_id = file_data['csv_file_id']
        csv_file_name = file_data['csv_file_name']
        pdf_file_id = file_data['pdf_file_id']
        pdf_file_name = file_data['pdf_file_name']
        store_code = file_data['store_code']

        processed_folder = create_folder_if_not_exists('Dominos Supplier Invoices - Processed', management_tenant_id, user)
        processed_folder_id = processed_folder.id


     
    

        # Fetch the content of the CSV file
        csv_content = fetch_file_content(user, management_tenant_id, csv_file_id)
        add_log(f"Fetched CSV content for {csv_file_name}", log_type="general", user_id=user_id)

        try:
            df = pd.read_csv(csv_content)
        except Exception as e:
            add_log(f"Error reading CSV: {csv_file_name}, {str(e)}", log_type="errors", user_id=user_id)
            continue

        expected_columns = [
            'Transaction type', 'Transaction date', 'Invoice number',
            'Order reference', 'Store', 'Total net', 'Total VAT',
            'Total gross', 'Nominal code', 'Nominal amount net',
            'VAT rate', 'Nominal code name', 'Net signed', 'VAT signed'
        ]

        # Check for missing columns
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            add_log(f"Missing columns in CSV: {csv_file_name}: {missing_columns}", log_type="errors", user_id=user_id)
            continue

        # Prepare line items for this CSV file
        line_items = []
        dates = df['Transaction date'].tolist() 
        store_number = df['Store'].iloc[0]
        total_invoice_amount = 0
        current_month = None
        previous_date = None
        

        # Iterate through each row in the DataFrame
        for index, row in df.iterrows():
            # Extract necessary fields
            description = f"{row['Invoice number']} {row['Nominal code name']}"
            unit_amount = float(row['Net signed']) if 'Net signed' in row and pd.notnull(row['Net signed']) else 0
            # Assign net_signed as Decimal, rounded to two decimal places
            net_signed = (
                Decimal(str(row['Net signed']))
                .quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            )

            # Assign vat_signed as Decimal, rounded to two decimal places
            vat_signed = (
                Decimal(str(row['VAT signed']))
                .quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            )
            unit_amount = unit_amount if unit_amount != 0 else float(row['Total gross']) if 'Total gross' in row and pd.notnull(row['Total gross']) else 0
            transaction_date_str = row['Transaction date']
            
            try:
                # Convert to datetime with flexible parsing
                transaction_date = pd.to_datetime(transaction_date_str, format='%Y/%m/%d', errors='coerce') # CHANGE FOR EMAIL UPLOAD FROM DOMINOS

            
                # Check if parsing failed (returns NaT if failed)
                if pd.isna(transaction_date):
                    add_log(f"Error parsing date '{transaction_date_str}' for tenant '{tenant_name}'", log_type="error", user_id=user_id)
                    continue  # Skip this row if the date is invalid

                
            except Exception as e:
                add_log(f"Error converting date: {str(e)}", log_type="error", user_id=user_id)
                continue

            statement_date = str(transaction_date.strftime('%d/%m/%Y'))

            # Extract month and year for grouping
            row_month = transaction_date.strftime('%Y-%m')  # Format: 'YYYY-MM'
            
            # Determine if we need to start a new invoice
            if current_month is None:
                current_month = row_month
                current_reference = "A"
                add_log(f"Starting new invoice for month: {current_month}", log_type="general", user_id=user_id)
            elif row_month != current_month:
                # Finalize the current invoice
                add_log(f"Finalizing invoice for month: {current_month} with {len(line_items)} line items", log_type="general", user_id=user_id)
                
                processed_files.append({
                    'tenant_id': tenant_id,
                    'tenant_name': tenant_name,
                    'line_items': line_items,
                    'csv_file_name': csv_file_name,
                    'csv_file_id': csv_file_id,
                    'pdf_file_id': pdf_file_id,
                    'pdf_file_name': pdf_file_name,
                    'dates': dates,
                    'store_number': store_number,
                    'total_invoice_amount': total_invoice_amount,
                    'processed_folder_id': processed_folder_id,
                    'store_code': store_code,
                    'statement_date': previous_date,
                    'reference': current_reference

                })
                
                # Reset for the new invoice
                line_items = []
                current_month = row_month
                current_reference = "B"  # Change to "B" for a new month
                total_invoice_amount = 0
                add_log(f"Starting new invoice for month: {current_month}", log_type="general", user_id=user_id)
            
            # Continue processing the current row
            vat_rate = float(row['VAT rate']) if 'VAT rate' in row and pd.notnull(row['VAT rate']) else 0
            tax_rate = 'ZERORATEDINPUT' if vat_rate == 0 else 'INPUT2' if vat_rate == 20 else None
            store_number = str(row['Store']).strip()
            
            # Get tracking info for the store
            tracking_info = tracking_codes.get(store_number)
            if tracking_info is None:
                add_log(f"Warning: No tracking category found for store {store_number}", log_type="error", user_id=user_id)
                continue
            
            # Retrieve the nominal code and account code for the user
            nominal_code = row['Nominal code']
            nominal_code_record = DomNominalCodes.query.filter_by(nominal_code=str(nominal_code), user_id=user_id).first()
            
            # Check if nominal code exists
            if nominal_code_record:
                # Check if store account code exists for the nominal code
                store_account_code = StoreAccountCodes.query.filter_by(id=nominal_code_record.store_account_code_id, user_id=user_id).first()
                if store_account_code:
                    account_code = store_account_code.account_code
                else:
                    # Log the error and raise an exception to stop the task
                    error_message = f"Store account code not found for nominal code {nominal_code} for user ID {user_id}."
                    add_log(error_message, log_type="errors", user_id=user_id)
                    task_status.status = 'failed'
                    task_status.result = error_message
                    db.session.commit()
                    raise Exception(error_message)  # Stop the task with an error
            else:
                # Log the error and raise an exception to stop the task
                error_message = f"Nominal code {nominal_code} not found in the database for user ID {user_id}."
                add_log(error_message, log_type="errors", user_id=user_id)
                task_status.status = 'failed'
                task_status.result = error_message
                db.session.commit()
                raise Exception(error_message)  # Stop the task with an error
            
            # Create line item tracking
            line_item_tracking = LineItemTracking(
                tracking_category_id=tracking_info['tracking_category_id'],
                tracking_option_id=tracking_info['tracking_option_id']
            )
            
            # Create line item
            line_item = LineItem(
                description=description,
                quantity=1,
                unit_amount=unit_amount,
                account_code=str(account_code),
                tax_type=tax_rate,
                tracking=[line_item_tracking]
            )
            
            line_items.append(line_item)

            total_invoice_amount = total_invoice_amount + vat_signed + net_signed
            
            previous_date = statement_date

        
        # After iterating, finalize the last invoice if there are remaining line items
        if line_items:
            add_log(f"Finalizing invoice for month: {current_month} with {len(line_items)} line items", log_type="general", user_id=user_id)
            
            processed_files.append({
                'tenant_id': tenant_id,
                'tenant_name': tenant_name,
                'line_items': line_items,
                'csv_file_name': csv_file_name,
                'csv_file_id': csv_file_id,
                'pdf_file_id': pdf_file_id,
                'pdf_file_name': pdf_file_name,
                'dates': dates,
                'store_number': store_number,
                'total_invoice_amount': total_invoice_amount,
                'processed_folder_id': processed_folder_id,
                'store_code': store_code,
                'statement_date': statement_date,
                'reference': current_reference
            })
        
        add_log(f"Processed {len(processed_files)} invoices from CSV {csv_file_name}", log_type="general", user_id=user_id) 

    # Group the processed files by tenant_id
    tenant_invoices = defaultdict(list)
    for file_data in processed_files:
        tenant_invoices[file_data['tenant_id']].append(file_data)

    
    

    # Loop through each tenant and process their invoices
    for tenant_id, files in tenant_invoices.items():
        tenant_name = files[0]['tenant_name']  # All files for the same tenant will have the same tenant_name
        invoices_to_create = []

        # Loop through each file for this tenant
        for file_data in files:
            csv_file_name = file_data['csv_file_name']
            pdf_file_name = file_data['pdf_file_name']
            csv_file_id = file_data['csv_file_id']
            pdf_file_id = file_data['pdf_file_id']
            line_items = file_data['line_items']
            dates = file_data['dates']  # Dates extracted from 'Transaction date' column
            store_number = file_data['store_number']  # Extracted from 'Store' column in the CSV file
            total_invoice_amount = file_data['total_invoice_amount']
            processed_folder_id = file_data['processed_folder_id']
            statement_date = file_data['statement_date']
            reference = file_data['reference']
            
            
            # Create invoice structure for this file
            invoice = {
                'tenant_id': tenant_id,
                'tenant_name': tenant_name,
                'invoice_date': dates[0],  # Assuming you want to use the first date as the invoice date
                'line_items': line_items,
                'csv_file_name': csv_file_name,
                'pdf_file_name': pdf_file_name,
                'csv_file_id': csv_file_id,
                'pdf_file_id': pdf_file_id,
                'store_number': store_number,
                'total_invoice_amount': total_invoice_amount,
                'processed_folder_id': processed_folder_id,
                'store_code': store_code,
                'statement_date': statement_date,
                'reference': reference 
    
            }

            invoices_to_create.append(invoice)
        
        # Bulk import these invoices for this tenant
        add_log(f"Preparing to bulk import {len(invoices_to_create)} invoices for tenant {tenant_name} ({tenant_id})", log_type="general", user_id=user_id)
        
        # This is where you'd call the actual function to bulk import to Xero or your system
        try:
            # Assuming you have a function `bulk_create_invoices` for bulk invoice creation

            response = bulk_create_bills(user, tenant_id, invoices_to_create, management_tenant_id)

            processed_file_count = processed_file_count + len(invoices_to_create)
            self.update_state(state='PROGRESS', meta={'current': processed_file_count, 'total': total_files})
            
            if response:  # Proceed only if the bulk_create_bills call was successful
                for invoice in invoices_to_create: 
                    add_log(f"Tenant Name '{tenant_name}' Processed dom purchase invoice '{invoice['csv_file_name']}' with attached pdf '{invoice['pdf_file_name']}', Total invoice amount '{invoice['total_invoice_amount']}', Statement Date '{invoice['statement_date']}', Store code '{invoice['store_number']}'", log_type="dom_purchase_invoice", user_id=user_id)

                    # Move and rename the CSV file
                    new_csv_name = f"{os.path.splitext(invoice['csv_file_name'])[0]}_PROCESSED.csv"
                    move_and_rename_file(invoice['csv_file_id'], invoice['processed_folder_id'], new_csv_name, "CSV", user, management_tenant_id)

                    # Move and rename the PDF file
                    new_pdf_name = f"{os.path.splitext(invoice['pdf_file_name'])[0]}_PROCESSED.pdf"
                    move_and_rename_file(invoice['pdf_file_id'], invoice['processed_folder_id'], new_pdf_name, "PDF", user, management_tenant_id)
                
                add_log(f"Successfully imported {len(invoices_to_create)} invoices for tenant {tenant_name} ({tenant_id})", log_type="general", user_id=user_id)
                
        except Exception as e:
            add_log(f"Failed to bulk import invoices for tenant {tenant_name} ({tenant_id}): {str(e)}", log_type="error", user_id=user_id)

    add_log(f"Dom purchase invoice processing completed for all tenants", log_type="general", user_id=user_id)


    task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

    if task_status:
        task_status.status = 'completed'
        task_status.result = "Task completed successfully."
        db.session.commit()


    self.update_state(state='SUCCESS', meta={'current': total_files, 'total': total_files})

    return {'current': total_files, 'total': total_files, 'status': 'Task completed!'}




    



@shared_task(bind=True, name='app.celery_tasks.process_dom_purchase_invoices_task_OLD')
def process_dom_purchase_invoices_task_OLD(self, user_id):
     # Load tracking codes from the database
    tracking_codes = {}
    categories = TrackingCategoryModel.query.filter_by(user_id=user_id).all()

    for category in categories:
        if category.store_number:  # Only add entries with a valid store number
            tracking_codes[category.store_number] = {
                "tracking_category_id": category.tracking_category_id,
                "tracking_option_id": category.tracking_option_id
            }
            
    add_log(f"Tracking codes loaded from the database", log_type="general", user_id=user_id)

    # Retrieve the user object from the database
    user = User.query.get(user_id)
    data = fetch_dom_invoicing_data(user)
    filtered_list = []

    total_files = sum([len(tenant['files']) for tenant in data])
    processed_file_count = 0



    # Prepare a list to store extracted details from files
    extracted_files = []
    processed_file_names = set()  # Track processed file names

    # Loop over each tenant
    for tenant in data:
        tenant_id = tenant['tenant_id']
        tenant_name = tenant['tenant_name']
        files = tenant['files']
        folders = tenant['folders']

        # Identify the Inbox folder ID
        inbox_folder = next((folder for folder in folders if folder['folder_name'].lower() == 'inbox'), None)
        processed_folder = create_folder_if_not_exists('Dominos Purchase Invoices - Processed', tenant_id, user)
        rejected_folder = create_folder_if_not_exists('Dominos Purchase Invoices -  Rejected', tenant_id, user)

        rejected_folder_id = rejected_folder.id
        processed_folder_id = processed_folder.id
        inbox_folder_id = inbox_folder['folder_id']

        # Filter files to only process those from the inbox
        inbox_files = [file for file in files if file['folder_id'] == inbox_folder_id and file['file_name'].startswith('CustAccountStatementExt.Report')]


        # Process each file
        for file in inbox_files:
            
            file_name = file['file_name']


            store_code_match = re.search(r'S-(\d+)-', file_name)
            store_code = store_code_match.group(1) if store_code_match else None

            # If the MIME type is incorrect, check the file extension
            mime_type = file['mime_type']
            if mime_type == 'application/octet-stream':
                if file_name.lower().endswith('.csv'):
                    mime_type = 'text/csv'
                elif file_name.lower().endswith('.pdf'):
                    mime_type = 'application/pdf'

            

            # If it's a PDF
            if mime_type == 'application/pdf' or file_name.lower().endswith('.pdf'):
                statement_date = extract_statement_date_from_pdf(tenant_id, file['file_id'],user)
                print("pdf: %s" % statement_date)


            # If it's a CSV
            elif mime_type  == 'text/csv' or file_name.lower().endswith('.csv'):
                statement_date = extract_statement_date_from_csv(tenant_id, file['file_id'], user)
                print("csv: %s" % statement_date)
                

            # Now we need to check if this file has already been processed
            if store_code and statement_date:
                # Remove both .csv and .pdf extensions from the file name
                base_file_name = file_name.replace('.csv', '').replace('.pdf', '')

                processed_log = LogEntry.query.filter(
                    and_(
                        LogEntry.message.like(f"%Store code '{store_code}'%"),
                        LogEntry.message.like(f"%Statement Date '{statement_date}'%"),
                        LogEntry.log_type == "dom_purchase_invoice",
                        LogEntry.user_id == user_id
                    )
                ).first()

            
                # If the file has already been processed or is a duplicate, move it to rejected
                if processed_log or file_name in processed_file_names:
                    add_log(f"File {file_name} (Store: {store_code}, Statement Date: {statement_date}) has already been processed or is a duplicate.", log_type="general", user_id=user_id)
                    rejected_filename = f"{os.path.splitext(file_name)[0]} (Rejected as duplicate or already processed){os.path.splitext(file_name)[1]}"

                    # Move and rename the CSV or PDF file to the "Rejected" folder
                    move_and_rename_file(file['file_id'], rejected_folder_id, rejected_filename, mime_type .split('/')[-1].upper(), user, tenant_id)
                    continue  # Skip further processing for this file

                # If not already processed, add to extracted list
                extracted_files.append({
                    'tenant_id': tenant_id,
                    'tenant_name': tenant_name,
                    'file_id': file['file_id'],
                    'file_name': file_name,
                    'store_code': store_code,
                    'statement_date': statement_date,
                    'mime_type': mime_type,
                    'processed_folder_id': processed_folder_id,
                })

                # Add to processed file names set
                processed_file_names.add(file_name)
        

    # Now that we have all the extracted details, let's match the CSV and PDF files
    matched_store_dates = set()  # Track matched store codes and dates to identify duplicates

    for csv_file in [f for f in extracted_files if f['mime_type'] == 'text/csv']:
        # Find the matching PDF file based on store_code and statement_date
        matching_pdf = next((f for f in extracted_files if f['mime_type'] == 'application/pdf' and f['store_code'] == csv_file['store_code'] and f['statement_date'] == csv_file['statement_date']), None)

        store_date_key = (csv_file['store_code'], csv_file['statement_date'])

        if store_date_key in matched_store_dates:
            # Duplicate store number and statement date, move both CSV and PDF to rejected
            add_log(f"Duplicate files found for Store {csv_file['store_code']} on {csv_file['statement_date']}. Moving to Rejected.", log_type="general", user_id=user_id)

            # Move the CSV to rejected
            rejected_csv_filename = f"{os.path.splitext(csv_file['file_name'])[0]} (Rejected as duplicate){os.path.splitext(csv_file['file_name'])[1]}"
            move_and_rename_file(csv_file['file_id'], rejected_folder_id, rejected_csv_filename, "CSV", user, csv_file['tenant_id'])

            # Move the matching PDF to rejected if it exists
            if matching_pdf:
                rejected_pdf_filename = f"{os.path.splitext(matching_pdf['file_name'])[0]} (Rejected as duplicate){os.path.splitext(matching_pdf['file_name'])[1]}"
                move_and_rename_file(matching_pdf['file_id'], rejected_folder_id, rejected_pdf_filename, "PDF", user, matching_pdf['tenant_id'])

            continue  # Skip further processing for this file

        # If no duplicates found, add the matched files to the processed list
        matched_store_dates.add(store_date_key)

        if matching_pdf:
            # Add the CSV and matching PDF file info to the filtered list, along with statement date and store code
            filtered_list.append({
                'tenant_id': csv_file['tenant_id'],
                'tenant_name': csv_file['tenant_name'],
                'csv_file_id': csv_file['file_id'],
                'csv_file_name': csv_file['file_name'],
                'pdf_file_id': matching_pdf['file_id'],
                'pdf_file_name': matching_pdf['file_name'],
                'store_code': csv_file['store_code'],            # Add store code to the filtered list
                'statement_date': csv_file['statement_date'],    # Add statement date to the filtered list
                'processed_folder_id': csv_file['processed_folder_id'],
            })
            add_log(f"Matched CSV {csv_file['file_name']} with PDF {matching_pdf['file_name']} for Store {csv_file['store_code']} on {csv_file['statement_date']}", log_type="general", user_id=user_id)
        else:
            # Log if no matching PDF was found
            add_log(f"No matching PDF found for CSV {csv_file['file_name']} (Store {csv_file['store_code']} on {csv_file['statement_date']})", log_type="error", user_id=user_id)
    


            

    total_files = len(filtered_list)

    # Create a list to store the processed data
    processed_files = []

    # Initialize tracking variables
    current_month = None

    current_reference = "A"  # Start with "A"

    # Iterate over the filtered files
    for file_data in filtered_list:
        tenant_id = file_data['tenant_id']
        tenant_name = file_data['tenant_name']
        csv_file_id = file_data['csv_file_id']
        csv_file_name = file_data['csv_file_name']
        pdf_file_id = file_data['pdf_file_id']
        pdf_file_name = file_data['pdf_file_name']
        processed_folder_id = file_data['processed_folder_id']
        store_code = file_data['store_code']
    

        # Fetch the content of the CSV file
        csv_content = fetch_file_content(user, tenant_id, csv_file_id)
        add_log(f"Fetched CSV content for {csv_file_name}", log_type="general", user_id=user_id)

        try:
            df = pd.read_csv(csv_content)
        except Exception as e:
            add_log(f"Error reading CSV: {csv_file_name}, {str(e)}", log_type="errors", user_id=user_id)
            continue

        expected_columns = [
            'Transaction type', 'Transaction date', 'Invoice number',
            'Order reference', 'Store', 'Total net', 'Total VAT',
            'Total gross', 'Nominal code', 'Nominal amount net',
            'VAT rate', 'Nominal code name', 'Net signed', 'VAT signed'
        ]

        # Check for missing columns
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            add_log(f"Missing columns in CSV: {csv_file_name}: {missing_columns}", log_type="errors", user_id=user_id)
            continue

        # Prepare line items for this CSV file
        line_items = []
        dates = df['Transaction date'].tolist() 
        store_number = df['Store'].iloc[0]
        total_invoice_amount = 0
        current_month = None
        previous_date = None
        

        # Iterate through each row in the DataFrame
        for index, row in df.iterrows():
            # Extract necessary fields
            description = f"{row['Invoice number']} {row['Nominal code name']}"
            unit_amount = float(row['Net signed']) if 'Net signed' in row and pd.notnull(row['Net signed']) else 0
            # Assign net_signed as Decimal, rounded to two decimal places
            net_signed = (
                Decimal(str(row['Net signed']))
                .quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            )

            # Assign vat_signed as Decimal, rounded to two decimal places
            vat_signed = (
                Decimal(str(row['VAT signed']))
                .quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            )
            unit_amount = unit_amount if unit_amount != 0 else float(row['Total gross']) if 'Total gross' in row and pd.notnull(row['Total gross']) else 0
            transaction_date_str = row['Transaction date']
            
            try:
                # Convert to datetime with flexible parsing
                transaction_date = pd.to_datetime(transaction_date_str, format='%Y/%m/%d', errors='coerce') # CHANGE FOR EMAIL UPLOAD FROM DOMINOS

            
                # Check if parsing failed (returns NaT if failed)
                if pd.isna(transaction_date):
                    add_log(f"Error parsing date '{transaction_date_str}' for tenant '{tenant_name}'", log_type="error", user_id=user_id)
                    continue  # Skip this row if the date is invalid

                
            except Exception as e:
                add_log(f"Error converting date: {str(e)}", log_type="error", user_id=user_id)
                continue

            statement_date = str(transaction_date.strftime('%d/%m/%Y'))

            # Extract month and year for grouping
            row_month = transaction_date.strftime('%Y-%m')  # Format: 'YYYY-MM'
            
            # Determine if we need to start a new invoice
            if current_month is None:
                current_month = row_month
                current_reference = "A"
                add_log(f"Starting new invoice for month: {current_month}", log_type="general", user_id=user_id)
            elif row_month != current_month:
                # Finalize the current invoice
                add_log(f"Finalizing invoice for month: {current_month} with {len(line_items)} line items", log_type="general", user_id=user_id)
                
                processed_files.append({
                    'tenant_id': tenant_id,
                    'tenant_name': tenant_name,
                    'line_items': line_items,
                    'csv_file_name': csv_file_name,
                    'csv_file_id': csv_file_id,
                    'pdf_file_id': pdf_file_id,
                    'pdf_file_name': pdf_file_name,
                    'dates': dates,
                    'store_number': store_number,
                    'total_invoice_amount': total_invoice_amount,
                    'processed_folder_id': processed_folder_id,
                    'store_code': store_code,
                    'statement_date': previous_date,
                    'reference': current_reference

                })
                
                # Reset for the new invoice
                line_items = []
                current_month = row_month
                current_reference = "B"  # Change to "B" for a new month
                total_invoice_amount = 0
                add_log(f"Starting new invoice for month: {current_month}", log_type="general", user_id=user_id)
            
            # Continue processing the current row
            vat_rate = float(row['VAT rate']) if 'VAT rate' in row and pd.notnull(row['VAT rate']) else 0
            tax_rate = 'ZERORATEDINPUT' if vat_rate == 0 else 'INPUT2' if vat_rate == 20 else None
            store_number = str(row['Store']).strip()
            
            # Get tracking info for the store
            tracking_info = tracking_codes.get(store_number)
            if tracking_info is None:
                add_log(f"Warning: No tracking category found for store {store_number}", log_type="error", user_id=user_id)
                continue
            
            # Retrieve the nominal code and account code for the user
            nominal_code = row['Nominal code']
            nominal_code_record = DomNominalCodes.query.filter_by(nominal_code=str(nominal_code), user_id=user_id).first()
            
            # Check if nominal code exists
            if nominal_code_record:
                # Check if store account code exists for the nominal code
                store_account_code = StoreAccountCodes.query.filter_by(id=nominal_code_record.store_account_code_id, user_id=user_id).first()
                if store_account_code:
                    account_code = store_account_code.account_code
                else:
                    # Log the error and raise an exception to stop the task
                    error_message = f"Store account code not found for nominal code {nominal_code} for user ID {user_id}."
                    add_log(error_message, log_type="errors", user_id=user_id)
                    raise Exception(error_message)  # Stop the task with an error
            else:
                # Log the error and raise an exception to stop the task
                error_message = f"Nominal code {nominal_code} not found in the database for user ID {user_id}."
                add_log(error_message, log_type="errors", user_id=user_id)
                raise Exception(error_message)  # Stop the task with an error
            
            # Create line item tracking
            line_item_tracking = LineItemTracking(
                tracking_category_id=tracking_info['tracking_category_id'],
                tracking_option_id=tracking_info['tracking_option_id']
            )
            
            # Create line item
            line_item = LineItem(
                description=description,
                quantity=1,
                unit_amount=unit_amount,
                account_code=str(account_code),
                tax_type=tax_rate,
                tracking=[line_item_tracking]
            )
            
            line_items.append(line_item)

            total_invoice_amount = total_invoice_amount + vat_signed + net_signed
            
            previous_date = statement_date

        
        # After iterating, finalize the last invoice if there are remaining line items
        if line_items:
            add_log(f"Finalizing invoice for month: {current_month} with {len(line_items)} line items", log_type="general", user_id=user_id)
            
            processed_files.append({
                'tenant_id': tenant_id,
                'tenant_name': tenant_name,
                'line_items': line_items,
                'csv_file_name': csv_file_name,
                'csv_file_id': csv_file_id,
                'pdf_file_id': pdf_file_id,
                'pdf_file_name': pdf_file_name,
                'dates': dates,
                'store_number': store_number,
                'total_invoice_amount': total_invoice_amount,
                'processed_folder_id': processed_folder_id,
                'store_code': store_code,
                'statement_date': statement_date,
                'reference': current_reference
            })
        
        add_log(f"Processed {len(processed_files)} invoices from CSV {csv_file_name}", log_type="general", user_id=user_id) 

    # Group the processed files by tenant_id
    tenant_invoices = defaultdict(list)
    for file_data in processed_files:
        tenant_invoices[file_data['tenant_id']].append(file_data)

    # Loop through each tenant and process their invoices
    for tenant_id, files in tenant_invoices.items():
        tenant_name = files[0]['tenant_name']  # All files for the same tenant will have the same tenant_name
        invoices_to_create = []

        # Loop through each file for this tenant
        for file_data in files:
            csv_file_name = file_data['csv_file_name']
            pdf_file_name = file_data['pdf_file_name']
            csv_file_id = file_data['csv_file_id']
            pdf_file_id = file_data['pdf_file_id']
            line_items = file_data['line_items']
            dates = file_data['dates']  # Dates extracted from 'Transaction date' column
            store_number = file_data['store_number']  # Extracted from 'Store' column in the CSV file
            total_invoice_amount = file_data['total_invoice_amount']
            processed_folder_id = file_data['processed_folder_id']
            statement_date = file_data['statement_date']
            reference = file_data['reference']
            
            
            # Create invoice structure for this file
            invoice = {
                'tenant_id': tenant_id,
                'tenant_name': tenant_name,
                'invoice_date': dates[0],  # Assuming you want to use the first date as the invoice date
                'line_items': line_items,
                'csv_file_name': csv_file_name,
                'pdf_file_name': pdf_file_name,
                'csv_file_id': csv_file_id,
                'pdf_file_id': pdf_file_id,
                'store_number': store_number,
                'total_invoice_amount': total_invoice_amount,
                'processed_folder_id': processed_folder_id,
                'store_code': store_code,
                'statement_date': statement_date,
                'reference': reference 
    
            }

            invoices_to_create.append(invoice)
        
        # Bulk import these invoices for this tenant
        add_log(f"Preparing to bulk import {len(invoices_to_create)} invoices for tenant {tenant_name} ({tenant_id})", log_type="general", user_id=user_id)
        
        # This is where you'd call the actual function to bulk import to Xero or your system
        try:
            # Assuming you have a function `bulk_create_invoices` for bulk invoice creation

            response = bulk_create_bills(user, tenant_id, invoices_to_create)

            processed_file_count = processed_file_count + len(invoices_to_create)
            self.update_state(state='PROGRESS', meta={'current': processed_file_count, 'total': total_files})
            
            if response:  # Proceed only if the bulk_create_bills call was successful

                for invoice in invoices_to_create: 
                    add_log(f"Tenant Name '{tenant_name}' Processed dom purchase invoice '{invoice['csv_file_name']}' with attached pdf '{invoice['pdf_file_name']}', Total invoice amount '{invoice['total_invoice_amount']}', Statement Date '{invoice['statement_date']}', Store code '{invoice['store_number']}'", log_type="dom_purchase_invoice", user_id=user_id)

                    # Move and rename the CSV file
                    new_csv_name = f"{os.path.splitext(invoice['csv_file_name'])[0]}_PROCESSED.csv"
                    move_and_rename_file(invoice['csv_file_id'], invoice['processed_folder_id'], new_csv_name, "CSV", user, invoice['tenant_id'])

                    # Move and rename the PDF file
                    new_pdf_name = f"{os.path.splitext(invoice['pdf_file_name'])[0]}_PROCESSED.pdf"
                    move_and_rename_file(invoice['pdf_file_id'], invoice['processed_folder_id'], new_pdf_name, "PDF", user, invoice['tenant_id'])
                
                add_log(f"Successfully imported {len(invoices_to_create)} invoices for tenant {tenant_name} ({tenant_id})", log_type="general", user_id=user_id)
                
        except Exception as e:
            add_log(f"Failed to bulk import invoices for tenant {tenant_name} ({tenant_id}): {str(e)}", log_type="error", user_id=user_id)

    add_log(f"Dom purchase invoice processing completed for all tenants", log_type="general", user_id=user_id)

    self.update_state(state='SUCCESS', meta={'current': total_files, 'total': total_files})
    return {'current': total_files, 'total': total_files, 'status': 'Task completed!'}





@shared_task(bind=True)
def process_dom_sales_invoices_task(self, user_id):

    # Fetch the user from the database
    user = User.query.get(user_id)

    # Fetch the TaskStatus record for the current task
    task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

    if task_status:
        task_status.status = 'in_progress'
        task_status.result = "task_started"
        db.session.commit()

    # Fetch the data for the current user's company (including tenant info)
    tenant_data = fetch_dom_management_company_data(user)

    if not tenant_data:
        add_log(f"No tenant data found for user {user_id}.", log_type="error")
        return "No tenant data found."
    
    tracking_codes = {}
    categories = TrackingCategoryModel.query.filter_by(user_id=user_id).all()

    for category in categories:
        if category.store_number:  # Only add entries with a valid store number
            tracking_codes[category.store_number] = {
                "tracking_category_id": category.tracking_category_id,
                "tracking_option_id": category.tracking_option_id
            }
            
    add_log(f"Tracking codes loaded from the database", log_type="general", user_id=user_id)

    # Since there's only one tenant, extract it directly
    tenant = tenant_data[0]
    tenant_id = tenant['tenant_id']

    # Get or create the required folders: Inbox, Processed, Rejected
    inbox_folder = create_folder_if_not_exists('Inbox', tenant_id, user)
    processed_folder = create_folder_if_not_exists('Processed', tenant_id, user)
    rejected_folder = create_folder_if_not_exists('Rejected', tenant_id, user)

    if not inbox_folder or not processed_folder or not rejected_folder:
        add_log(f"Failed to retrieve or create one of the necessary folders for tenant {tenant_id}.", log_type="error", user_id=user_id)
        return "Folder creation failed."

    # Process files in the inbox folder
    inbox_files = [file for file in tenant['files'] if file['folder_id'] == inbox_folder.id]

    tenant_contact_details = get_all_contacts(user)
    
    # Map tenant names to contact IDs
    try:
        tenant_to_contact_mapping = map_tenant_to_contact(tenant_contact_details, user_id)
    except Exception as e:
        # If there's an error in finding contacts, log the error and stop the task
        add_log(f"Error in finding contacts: {str(e)}", log_type="error", user_id=user_id)
        error_message = str(e)

        task_status.status = 'failed'
        task_status.result = error_message
        db.session.commit()

        return f"Task unsuccessful: {str(e)}"
    

    for file in inbox_files:
        if file['file_name'].startswith('KeyIndicatorsStore'):

            # If the file is a dictionary, extract the file name
            if isinstance(file, dict) and 'file_name' in file:
                original_file_name = file['file_name']

                # Fetch the file content (assuming you are fetching it from Xero API or similar)
                temp_file_path = fetch_file_content(user, tenant_id, file['file_id'])

                # Ensure the file content is processed correctly (this depends on how you're handling the content)
                with open(temp_file_path, 'rb') as f:
                    complete_file_data = f.read()  # Read file content for attaching it later
                

                print(f"Processing file: {original_file_name}")

            else:
                print(f"Unexpected file object type: {type(file)}")

            # Read the data from A4 and A5
            file_info = pd.read_excel(temp_file_path, sheet_name=0, skiprows=2, nrows=2, usecols="A")

            # Initialize variables to hold the start and end dates
            start_date = None
            end_date = None

            # Use regular expressions to extract the dates
            for index, row in file_info.iterrows():
                cell_value = row.iloc[0]  # Get the content of the cell

                if isinstance(cell_value, str):
                    # Look for the "Start Date" pattern
                    match_start = re.search(r"Start Date\s*:\s*(\d{2}-[A-Za-z]{3}-\d{4})", cell_value)
                    if match_start:
                        start_date = match_start.group(1)

                    # Look for the "End Date" pattern
                    match_end = re.search(r"End Date\s*:\s*(\d{2}-[A-Za-z]{3}-\d{4})", cell_value)
                    if match_end:
                        end_date = match_end.group(1)


            # Check if the file has already been processed by matching file name, start date, and end date
            processed_log = LogEntry.query.filter(
                and_(
                    LogEntry.message.like(f"%Processed dom sales invoice '{file['file_name']}'%"),
                    LogEntry.message.like(f"%Start Date '{start_date}'%"),
                    LogEntry.message.like(f"%End Date '{end_date}'%"),
                    LogEntry.log_type == "dom_sales_invoice",
                    LogEntry.user_id == user_id
                )
            ).first()


            if processed_log:
                tenant_id = tenant['tenant_id']
                # Log duplicate and move to rejected folder
                add_log(f"File {file['file_name']} has already been processed.", log_type="general", user_id=user.id)
                rejected_filename = f"{os.path.splitext(file['file_name'])[0]} (Rejected as Duplicate){os.path.splitext(file['file_name'])[1]}"

                # Use the move_and_rename_file function to move and rename the file
                move_and_rename_file(
                    file_id=file['file_id'],
                    new_folder_id=rejected_folder.id,
                    new_name=rejected_filename,
                    file_type="Duplicate",
                    user=user,
                    tenant_id=tenant_id
                )
                continue


            month = start_date.split('-')[1].upper()

            # Assuming start_date is a datetime object
            mileage_description = f"Mileage for {start_date} to {end_date}"
            sales_description = f"Sales for {start_date} to {end_date}"


            # Load Excel file and process data
            df = pd.read_excel(temp_file_path, sheet_name=0, skiprows=6, usecols="A,F,G,AA")
            df['Unit price for 20% VAT'] = (df['Total\nTax'] / 0.2)
            df['Unit price for Zero Rated'] = df['Total\nSales'] - (df['Total\nTax'] / 0.2) - df['Total\nTax']
            df['Unit price for 20% VAT for Mileage'] = (df['Delivered\nCost'] * 0.9) / 1.2
            df['Unit price for Zero Rated for Mileage'] = df['Delivered\nCost'] * 0.1
            df2 = pd.read_excel(temp_file_path, sheet_name=0, skiprows=3, nrows=1, usecols="A")
            date = df2.iloc[0, 0]

            line_items_sales = []
            line_items_mileage = []

            missing_store_errors = []

            # Create sales and mileage line items
            for _, row in df.iterrows():
                store_number = str(int(row['Store\nID'])).strip()
                tracking_info = tracking_codes.get(store_number)
                if not tracking_info:
                    missing_store_errors.append(f"Missing tracking category for store number '{store_number}' in file '{original_file_name}'")
                    continue

                line_item_tracking = LineItemTracking(
                    tracking_category_id=tracking_info['tracking_category_id'],
                    tracking_option_id=tracking_info['tracking_option_id']
                )

                # Sales line item for 20% VAT
                line_items_sales.append(LineItem(
                    description=sales_description,
                    quantity=1,
                    unit_amount=str(row['Unit price for 20% VAT']),
                    account_code='4001',
                    tax_type="OUTPUT2",
                    tracking=[line_item_tracking]
                ))

                # Sales line item for Zero Rated
                line_items_sales.append(LineItem(
                    description=sales_description,
                    quantity=1,
                    unit_amount=str(row['Unit price for Zero Rated']),
                    account_code='4001',
                    tax_type="ZERORATEDOUTPUT",
                    tracking=[line_item_tracking]
                ))

                # Mileage line items for 20% VAT and Zero Rated
                line_items_mileage.append(LineItem(
                    description=mileage_description,
                    quantity=1,
                    unit_amount=str(row['Unit price for 20% VAT for Mileage']),
                    account_code='7302',
                    tax_type="INPUT2",
                    tracking=[line_item_tracking]
                ))

                line_items_mileage.append(LineItem(
                    description=mileage_description,
                    quantity=1,
                    unit_amount=str(row['Unit price for Zero Rated for Mileage']),
                    account_code='7302',
                    tax_type="ZERORATEDINPUT",
                    tracking=[line_item_tracking]
                ))

            # Check if there were any missing store numbers and raise an exception
            if missing_store_errors:
                # Join the missing store numbers into a string for logging and renaming
                error_message = "\n".join(missing_store_errors)
                missing_store_numbers_str = ', '.join([msg.split("'")[1] for msg in missing_store_errors])  # Extract store numbers

                # Log the error with missing store numbers
                add_log(f"Task failed due to missing store numbers: \n{error_message}", log_type="error", user_id=user_id)

                # Move the file to the "Rejected" folder and rename it to indicate missing store numbers
                rejected_filename = f"{os.path.splitext(original_file_name)[0]} (Rejected - Missing Store Numbers {missing_store_numbers_str}){os.path.splitext(original_file_name)[1]}"

                try:
                    move_and_rename_file(
                        file_id=file['file_id'],
                        new_folder_id=rejected_folder.id,
                        new_name=rejected_filename,
                        file_type="Rejected",
                        user=user,
                        tenant_id=tenant_id
                    )
                    add_log(f"File '{original_file_name}' moved to Rejected folder with new name '{rejected_filename}' due to missing store numbers.", log_type="general", user_id=user_id)
                except Exception as e:
                    add_log(f"Error moving file '{original_file_name}' to Rejected folder: {str(e)}", log_type="error", user_id=user_id)

                # Raise an exception to stop the task
                task_status.status = 'failed'
                task_status.result = error_message
                db.session.commit()

                raise Exception(f"Missing store numbers found. Task failed. \n{error_message}")

            # Group sales and mileage line items by tracking_option_id
            
            grouped_sales_items = defaultdict(list)
            grouped_mileage_items = defaultdict(list)

            for item in line_items_sales:
                tracking_option_id = item.tracking[0].tracking_option_id
                grouped_sales_items[tracking_option_id].append(item)

            for item in line_items_mileage:
                tracking_option_id = item.tracking[0].tracking_option_id
                grouped_mileage_items[tracking_option_id].append(item)


            # Fetch all tracking categories for the user
            tracking_categories = TrackingCategoryModel.query.filter_by(user_id=user_id).all()

            # Fetch all Xero tenants for the user
            xero_tenants = XeroTenant.query.filter_by(user_id=user_id).all()

            # Create a mapping of tenant_name to tenant_id using XeroTenant model
            tenant_name_to_id = {tenant.tenant_name: tenant.tenant_id for tenant in xero_tenants}

            # Create a mapping of tenant_id to tenant_name from the XeroTenant model
            tenant_id_to_name = {tenant.tenant_id: tenant.tenant_name for tenant in XeroTenant.query.filter_by(user_id=user_id).all()}

            # Create a mapping of tracking_option_id to tenant_id using tenant_name from TrackingCategoryModel
            tracking_category_to_tenant = {}
            for category in tracking_categories:
                tenant_id = tenant_name_to_id.get(category.tenant_name)
                if tenant_id:
                    tracking_category_to_tenant[category.tracking_option_id] = tenant_id
                else:
                    add_log(f"No tenant found for tenant name '{category.tenant_name}' in XeroTenant model.", log_type="error", user_id=user_id)


            # Post grouped sales line items to corresponding tenants
            for tracking_option_id, sales_items in grouped_sales_items.items():

                # Query to get the store name (tracking_category_option) based on tracking_option_id
                store_name = db.session.query(TrackingCategoryModel.tracking_category_option).filter_by(
                    tracking_option_id=tracking_option_id
                ).scalar()  # .scalar() fetches the single column value directly
    

                # Find the tenant_id associated with this tracking_option_id
                tenant_id = tracking_category_to_tenant.get(tracking_option_id)
                tenant_name = tenant_id_to_name.get(tenant_id)
                if not tenant_id:
                    add_log(f"No tenant found for tracking category {tracking_option_id}", log_type="error", user_id=user_id)
                    continue

                # Retrieve the contact IDs from the tenant-to-contact mapping
                tenant_contacts = tenant_to_contact_mapping.get(tenant_name, {})
                sales_contact_id = tenant_contacts.get("SALES")
                
                if not sales_contact_id:
                    add_log(f"Missing SALES contact for tenant {tenant_name}.", log_type="error", user_id=user_id)
                    continue

                # Post sales invoice
                try:
                    invoice_id_sales = post_dom_sales_invoice_with_attachment(
                        tenant_id, sales_items, sales_contact_id, original_file_name, complete_file_data, user, end_date, start_date, month, store_name, invoice_type="Domino's Sales",
                    )
                    add_log(f"Posted sales invoice {invoice_id_sales} for tenant {tenant_id} with tracking category {tracking_option_id}.", log_type="general", user_id=user_id)
                    add_log(f"Tenant Name '{tenant_name}' Processed dom sales invoice '{original_file_name}', Supplier type 'DOMINOS SALES', Start Date '{start_date}', End Date '{end_date}'", log_type="dom_sales_invoice", user_id=user_id)
                except Exception as e:
                    add_log(f"Error posting sales invoice for tenant {tenant_id} with tracking category {tracking_option_id}: {str(e)}", log_type="error", user_id=user_id)

            # Post grouped mileage line items to corresponding tenants
            for tracking_option_id, mileage_items in grouped_mileage_items.items():

                # Query to get the store name (tracking_category_option) based on tracking_option_id
                store_name = db.session.query(TrackingCategoryModel.tracking_category_option).filter_by(
                    tracking_option_id=tracking_option_id
                ).scalar()  # .scalar() fetches the single column value directly
    

                # Find the tenant_id associated with this tracking_option_id
                tenant_id = tracking_category_to_tenant.get(tracking_option_id)
                tenant_name = tenant_id_to_name.get(tenant_id)
                if not tenant_id:
                    add_log(f"No tenant found for tracking category {tracking_option_id}", log_type="error", user_id=user_id)
                    continue

                # Retrieve the contact IDs from the tenant-to-contact mapping
                tenant_contacts = tenant_to_contact_mapping.get(tenant_name, {})
                mileage_contact_id = tenant_contacts.get("MILEAGE")
                
                if not mileage_contact_id:
                    add_log(f"Missing MILEAGE contact for tenant {tenant_name}.", log_type="error", user_id=user_id)
                    continue

                # Post mileage invoice
                try:
                    invoice_id_mileage = post_dom_sales_invoice_with_attachment(
                        tenant_id, mileage_items, mileage_contact_id, original_file_name, complete_file_data, user, end_date, start_date, month, store_name, invoice_type="Domino's Mileage",
                    )
                    add_log(f"Posted mileage invoice {invoice_id_mileage} for tenant {tenant_id} with tracking category {tracking_option_id}.", log_type="general", user_id=user_id)
                    add_log(f"Tenant Name '{tenant_name}' Processed dom sales invoice '{original_file_name}', Supplier type 'DOMINOS MILEAGE', Start Date '{start_date}', End Date '{end_date}'", log_type="dom_sales_invoice", user_id=user_id)
                except Exception as e:
                    add_log(f"Error posting mileage invoice for tenant {tenant_id} with tracking category {tracking_option_id}: {str(e)}", log_type="error", user_id=user_id)


            

            # Post sales and mileage invoices to Xero
            # post_invoices_to_xero(api_client, tenant_id, line_items_sales, line_items_mileage, temp_file_path, file, date, user_id)

            # Move the file to the processed folder after posting
            try:
                tenant_id = tenant['tenant_id']
                rejected_filename = f"{os.path.splitext(file['file_name'])[0]} (PROCESSED){os.path.splitext(file['file_name'])[1]}"
                move_and_rename_file(
                    file_id=file['file_id'],
                    new_folder_id=processed_folder.id,
                    new_name=rejected_filename,
                    file_type="Sales Invoice Processed File",
                    user=user,
                    tenant_id=tenant_id
                )
                add_log(f"File {file['file_name']} moved to Processed folder.", log_type="general", user_id=user_id)
            except Exception as e:
                add_log(f"Error moving file {file['file_name']} to Processed folder: {str(e)}", log_type="error", user_id=user_id)

    task_status.status = 'completed'
    task_status.result = "Task completed successfully."
    db.session.commit()

    return "Processing complete"


def map_tenant_to_contact(tenant_contact_details, user_id):
    # Initialize a dictionary to store the mapping of tenant names to contact IDs
    tenant_to_contact_mapping = {}

    # Create a mapping of tenant names to their contact details
    tenant_name_to_contacts = {tenant['tenant_name']: tenant['contacts'] for tenant in tenant_contact_details}

    # Initialize a list to capture the error details for all missing contacts
    missing_contacts_info = []

    # Fetch tracking categories for the user, filter by tenants present in TrackingCategoryModel
    tracking_categories = TrackingCategoryModel.query.filter_by(user_id=user_id).all()

    # Get the tenant names from TrackingCategoryModel
    tracked_tenant_names = {category.tenant_name for category in tracking_categories}

    # Loop through tenant names that are in the tracking categories to map tenant to contact IDs
    for tenant_name, contacts in tenant_name_to_contacts.items():
        # Only proceed if this tenant exists in the TrackingCategoryModel
        if tenant_name not in tracked_tenant_names:
            continue
        
        # Initialize contact IDs for "DOMINOS MILEAGE" and "DOMINOS SALES"
        dominos_mileage_contact_id = None
        dominos_sales_contact_id = None

        # Loop through the contacts to find the matching contact names
        for contact in contacts:
            if contact['contact_name'] == "MILEAGE":
                dominos_mileage_contact_id = contact['contact_id']
            elif contact['contact_name'] == "SALES":
                dominos_sales_contact_id = contact['contact_id']

        # Check if both contacts were found
        if dominos_mileage_contact_id and dominos_sales_contact_id:
            # If both are found, map them to the tenant name
            tenant_to_contact_mapping[tenant_name] = {
                "MILEAGE": dominos_mileage_contact_id,
                "SALES": dominos_sales_contact_id
            }
        else:
            # If one or both contacts are missing, capture the details
            missing_contacts = []
            if not dominos_mileage_contact_id:
                missing_contacts.append("MILEAGE")
            if not dominos_sales_contact_id:
                missing_contacts.append("SALES")

            # Append the missing contact info for this tenant
            missing_contacts_info.append(
                f"Tenant '{tenant_name}' is missing: {', '.join(missing_contacts)}"
            )

    # If there are any missing contacts, log them all as one error
    if missing_contacts_info:
        # Join all missing contact messages into one error message
        error_message = "The following tenants are missing contacts:\n" + "\n".join(missing_contacts_info)
        add_log(error_message, log_type="error", user_id=user_id)
        raise Exception(error_message)

    return tenant_to_contact_mapping




def load_tracking_codes_by_store_postcodes(user_id):
    tracking_codes = {}
    
    try:
        # Ensure the user is logged in
        if user_id:
            # Query the database to get all tracking categories for the current user
            tracking_categories = TrackingCategoryModel.query.filter_by(user_id=user_id).all()

            # Loop through the results and populate the tracking_codes dictionary
            for category in tracking_categories:
                if category.store_postcode:  # Only add entries with a valid store postcode
                    tracking_codes[category.store_postcode] = {
                        "tenant_name": category.tenant_name,
                        "tracking_category_id": category.tracking_category_id,
                        "tracking_category_name": category.tracking_category_name,
                        "tracking_category_option": category.tracking_category_option,
                        "tracking_option_id": category.tracking_option_id
                    }
        else:
            add_log("User not logged in. No tracking codes can be loaded.", log_type="errors")
    except Exception as e:
        add_log(f"Error loading tracking codes from database: {str(e)}", log_type="errors")
    
    return tracking_codes



def extract_store_postcode(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            # Extract Ship-to & Sold-to Address
            ship_to_pattern = r"Ship-to & Sold-to\s*:\s*(.*?)\n\d{6,}"
            ship_to_match = re.search(ship_to_pattern, text, re.DOTALL)
            if ship_to_match:
                ship_to_address = ship_to_match.group(1).strip()
                
                # Extract the postcode from the address
                postcode_pattern = r"\b[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}\b"
                postcode_match = re.search(postcode_pattern, ship_to_address)
                
                if postcode_match:
                    postcode = postcode_match.group(0)
                else:
                    postcode = "Postcode not found"
            else:
                postcode = "Ship-to & Sold-to address not found"

            # Return only the postcode
            return {
                postcode
            }
    
    return {
        None
    }

def extract_store_number(pdf_path):
    # Updated pattern to match "S" followed by 5 digits and capture only the digits
    store_number_pattern = r"S(\d{5})"  # Captures the 5 digits after "S"
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            # Search for the store number pattern in the extracted text
            store_number_match = re.search(store_number_pattern, text)
            
            if store_number_match:
                # Extract only the digits (group 1) from the match
                store_number = store_number_match.group(1).strip()
                return store_number
            else:
                return "Store number not found"
    
    return None

def extract_postcode_from_delivery_address(pdf_path):
    # Pattern to capture a UK postcode (e.g., LA4 4DD)
    postcode_pattern = r"\b[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}\b"

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            # Search for "Delivery Address" and extract the text below it
            delivery_address_pattern = r"Delivery Address(.*?)(\b[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}\b)"
            delivery_address_match = re.search(delivery_address_pattern, text, re.DOTALL)
            
            if delivery_address_match:
                # Extract the postcode from the delivery address
                postcode_match = re.search(postcode_pattern, delivery_address_match.group(0))
                if postcode_match:
                    return postcode_match.group(0).strip()
                else:
                    return "Postcode not found"
    
    return None



def load_tracking_codes_by_store_number(user_id):
    tracking_codes = {}
    try:
        # Ensure the user is logged in
        if user_id:
            # Query the database to get all tracking categories for the current user based on store number
            tracking_categories = TrackingCategoryModel.query.filter_by(user_id).all()

            # Loop through the results and populate the tracking_codes dictionary
            for category in tracking_categories:
                if category.store_number:  # Only add entries with a valid store number
                    tracking_codes[category.store_number] = {
                        "tenant_name": category.tenant_name,
                        "tracking_category_id": category.tracking_category_id,
                        "tracking_category_name": category.tracking_category_name,
                        "tracking_category_option": category.tracking_category_option,
                        "tracking_option_id": category.tracking_option_id
                    }
        else:
            add_log("User not logged in. No tracking codes can be loaded.", log_type="errors")
    except Exception as e:
        add_log(f"Error loading tracking codes from database: {str(e)}", log_type="errors")
    
    return tracking_codes

def check_if_credit_memo(pdf_path):
    # Pattern to capture the phrase "CREDIT MEMO"
    credit_memo_pattern = r"CREDIT MEMO"

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            # Check for "CREDIT MEMO" in the text
            if re.search(credit_memo_pattern, text):
                print("Found 'CREDIT MEMO' in the PDF. Skipping processing.")
                return True  # It's a credit memo
            else:
                print("This is not a credit memo.")
    
    return False  # It's not a credit memo




@shared_task(bind=True)
def process_cocacola_task(self, user_id):
    # Fetch the user by user_id
    user = User.query.get(user_id)

    # Fetch all tenants associated with this user that are in Live Dom Tenants List
    tenants = DomPurchaseInvoicesTenant.query.filter_by(user_id=user_id).all()

    # Extract tenant names to pass into the invoice fetching function
    tenant_names = [tenant.tenant_name for tenant in tenants]

    # Fetch Coca-Cola invoices without tracking categories for these tenants
    cocacola_invoices= get_invoices_and_credit_notes(user, tenant_names, "Coca-Cola")




    # Calculate total invoices to process
    total_invoices = len(cocacola_invoices)
    processed_invoices = 0

    invoices_to_process = [invoice for invoice in cocacola_invoices if invoice['xero_type'] == "invoice"]
    credit_notes_to_process = [invoice for invoice in cocacola_invoices if invoice['xero_type'] == "credit_note"]

    # Track errors
    errors = []
    
    # Process the invoices for each tenant
    for invoice in invoices_to_process:

        print(invoice["invoice_reference"])
        print(invoice['invoice_id'])
        print(invoice["invoice_date"])

        data = extract_coca_cola_invoice_data(user, invoice)
        
        # Check if there are any errors
        invoice_errors = data.get("errors", [])

        if invoice_errors:
            errors.extend(invoice_errors)

            # Store invoice with errors in SupplierInvoiceRecord
            new_record = SupplierInvoiceRecord(
                user_id=user.id,
                store_name=invoice["tenant_name"],
                invoice_type="Coca-Cola",
                invoice_number=invoice["invoice_reference"],
                invoice_id=invoice['invoice_id'],
                errors=', '.join(invoice_errors),
                triggered_by="manual",  # Or "manual" if triggered manually
                date_of_invoice=invoice["invoice_date"]
            )
            db.session.add(new_record)
            db.session.commit()

            add_log(f"Error processing invoice {invoice['invoice_id']}: {', '.join(invoice_errors)}", log_type="error", user_id=user.id)
            continue  # Skip this invoice due to errors

        # Process based on invoice type
        if data["invoice_type"] == "credit_memo":
            # Call the function to convert the credit memo to an invoice
            error = convert_invoice_to_credit_memo(data, user)

        if data["invoice_type"] == "void":
            invoice_id = data['invoice_id']
            tenant_id = data['tenant_id']
            void_invoice(invoice_id, tenant_id, user)

        else:
            # Call the function to assign tracking code to the invoice
            error = assign_tracking_code_to_invoice(data, user)
        
        # Store invoice with errors in SupplierInvoiceRecord
        new_record = SupplierInvoiceRecord(
            user_id=user.id,
            store_name=invoice["tenant_name"],
            invoice_type="Coca-Cola",
            invoice_number=invoice["invoice_reference"],
            invoice_id=invoice['invoice_id'],
            triggered_by="manual",  # Or "manual" if triggered manually
            date_of_invoice=invoice["invoice_date"]
        )
        db.session.add(new_record)
        db.session.commit()

        # Update progress after processing each invoice
        processed_invoices += 1
        self.update_state(state='PROGRESS', meta={'current': processed_invoices, 'total': total_invoices})
    
    # Process the credit notes for each tenant
    for credit_note in credit_notes_to_process:

        data = extract_coca_cola_invoice_data(user, credit_note)

        # Check if there are any errors
        credit_note_errors = data.get("errors", [])
        if credit_note_errors:
            errors.extend(credit_note_errors)

            # Store invoice with errors in SupplierInvoiceRecord
            new_record = SupplierInvoiceRecord(
                user_id=user.id,
                store_name=credit_note["tenant_name"],
                invoice_type="Coca-Cola",
                invoice_number=credit_note["invoice_reference"],
                invoice_id=credit_note['invoice_id'],
                errors=', '.join(credit_note_errors),
                triggered_by="manual",  # Or "manual" if triggered manually
                date_of_invoice=credit_note["invoice_date"]
            )
            db.session.add(new_record)
            db.session.commit()

            add_log(f"Error processing credit note {credit_note['invoice_id']}: {', '.join(credit_note_errors)}", log_type="error", user_id=user.id)
            continue  # Skip this credit note due to errors


        # Process based on credit note type
        if data["invoice_type"] == "invoice":
            # Call the function to convert the credit memo to an invoice
            error = convert_credit_memo_to_invoice(data, user)
        else:
            # Call the function to assign tracking code to the credit memo
            error = assign_tracking_code_to_credit_note(data, user)
            add_log(f"Credit note {credit_note['invoice_id']} is a valid credit memo.", log_type="general", user_id=user.id)
        
        
        new_record = SupplierInvoiceRecord(
            user_id=user.id,
            store_name=credit_note["tenant_name"],
            invoice_type="Coca-Cola",
            invoice_number=credit_note["invoice_reference"],
            invoice_id=credit_note['invoice_id'],
            triggered_by="manual",  # Or "manual" if triggered manually
            date_of_invoice=credit_note["invoice_date"]
        )


        db.session.add(new_record)
        db.session.commit()


        # Update progress after processing each invoice
        processed_invoices += 1
        self.update_state(state='PROGRESS', meta={'current': processed_invoices, 'total': total_invoices})

    # Final message
    self.update_state(state='SUCCESS', meta={
        'current': total_invoices,
        'total': total_invoices,
        'message': "Coca-Cola invoice processing completed",
        'errors': errors  # Ensure errors are included in the meta response
    })

    # Step 1: Create a subquery to find the most recent record for each invoice_id
    latest_records_subquery = db.session.query(
        SupplierInvoiceRecord.invoice_id,
        func.max(SupplierInvoiceRecord.run_time).label("latest_run_time")
    ).filter(SupplierInvoiceRecord.invoice_id != None).group_by(SupplierInvoiceRecord.invoice_id).subquery()

    # Step 2: Create a subquery to find the IDs of older records
    subquery = db.session.query(
        SupplierInvoiceRecord.id
    ).join(
        latest_records_subquery,
        and_(
            SupplierInvoiceRecord.invoice_id == latest_records_subquery.c.invoice_id,
            SupplierInvoiceRecord.run_time < latest_records_subquery.c.latest_run_time
        )
    ).subquery()

    # Step 3: Use subquery.select() explicitly in the `.in_()` filter
    db.session.query(SupplierInvoiceRecord).filter(
        SupplierInvoiceRecord.id.in_(subquery.select())
    ).delete(synchronize_session=False)

    # Commit the changes
    db.session.commit()
    

    # Log the cleanup
    add_log("Removed duplicate entries from SupplierInvoiceRecord table, keeping only the most recent ones.", log_type="general", user_id=user.id)
        
    # Prepare final message
    if errors:
        return {
            "message": f"Coca-Cola invoice processing completed with {len(errors)} errors.",
            "errors": errors
        }
    else:
        return {"message": "Coca-Cola invoice processing completed successfully for all tenants."}



@shared_task(bind=True)
def process_textman_task(self, user_id):
    # Fetch the user by user_id
    user = User.query.get(user_id)
    
    # Fetch all tenants associated with this user from DomPurchaseInvoicesTenant
    tenants = DomPurchaseInvoicesTenant.query.filter_by(user_id=user_id).all()

    # Extract tenant names to pass into the invoice fetching function
    tenant_names = [tenant.tenant_name for tenant in tenants]

    # Fetch Textman invoices without tracking categories for these tenants
    textman_invoices = get_invoices_and_credit_notes(user, tenant_names, "Text Management")


    # Count total invoices
    total_invoices = len(textman_invoices)
    processed_invoices = 0

    # Track errors
    errors = []

    invoices_to_process = [invoice for invoice in textman_invoices if invoice['xero_type'] == "invoice"]
    credit_notes_to_process = [invoice for invoice in textman_invoices if invoice['xero_type'] == "credit_note"]

    # Process the invoices for each tenant
    for invoice in invoices_to_process:
        data = extract_textman_invoice_data(user, invoice)

        # Check if there are any errors
        invoice_errors = data.get("errors", [])
        if invoice_errors:
            errors.extend(invoice_errors)

            new_record = SupplierInvoiceRecord(
                user_id=user.id,
                store_name=invoice["tenant_name"],
                invoice_type="Text Management",
                invoice_number=invoice["invoice_reference"],
                invoice_id=invoice['invoice_id'],
                errors=', '.join(invoice_errors),
                triggered_by="manual",
                date_of_invoice=invoice["invoice_date"]
            )
            db.session.add(new_record)
            db.session.commit()

            add_log(f"Error processing Textman invoice {invoice['invoice_id']}: {', '.join(invoice_errors)}", log_type="error", user_id=user.id)
            continue  # Skip this invoice due to errors

        # Process based on invoice type
        if data["invoice_type"] == "credit_memo":
            # Call the function to convert the credit memo to an invoice
            error = convert_invoice_to_credit_memo(data, user)
        else:
            # Call the function to assign tracking code to the invoice
            error = assign_tracking_code_to_invoice(data, user)

        new_record = SupplierInvoiceRecord(
            user_id=user.id,
            store_name=invoice["tenant_name"],
            invoice_type="Text Management",
            invoice_number=invoice["invoice_reference"],
            invoice_id=invoice['invoice_id'],
            triggered_by="manual",
            date_of_invoice=invoice["invoice_date"]
        )
        db.session.add(new_record)
        db.session.commit()

        # Log progress for this invoice
        processed_invoices += 1
        self.update_state(state='PROGRESS', meta={'current': processed_invoices, 'total': total_invoices})

    # Process the credit notes for each tenant
    for credit_note in credit_notes_to_process:
        data = extract_textman_invoice_data(user, credit_note)

        # Check if there are any errors
        credit_note_errors = data.get("errors", [])
        if credit_note_errors:

            errors.extend(credit_note_errors)

            new_record = SupplierInvoiceRecord(
                user_id=user.id,
                store_name=credit_note["tenant_name"],
                invoice_type="Text Management",
                invoice_number=credit_note["invoice_reference"],
                invoice_id=credit_note['invoice_id'],
                errors=', '.join(credit_note_errors),
                triggered_by="manual",
                date_of_invoice=credit_note["invoice_date"]
            )
            db.session.add(new_record)
            db.session.commit()

            add_log(f"Error processing Textman credit note {credit_note['invoice_id']}: {', '.join(credit_note_errors)}", log_type="error", user_id=user.id)
            continue  # Skip this credit note due to errors

        # Process based on credit note type
        if data["invoice_type"] == "invoice":
            # Call the function to convert the credit memo to an invoice
            error = convert_credit_memo_to_invoice(data, user)
        else:
            # Call the function to assign tracking code to the credit note
            error = assign_tracking_code_to_credit_note(data, user)

    

        new_record = SupplierInvoiceRecord(
            user_id=user.id,
            store_name=credit_note["tenant_name"],
            invoice_type="Text Management",
            invoice_number=credit_note["invoice_reference"],
            invoice_id=credit_note['invoice_id'],
            triggered_by="manual",
            date_of_invoice=credit_note["invoice_date"]
        )
        db.session.add(new_record)
        db.session.commit()

        # Log progress for this credit note
        processed_invoices += 1
        self.update_state(state='PROGRESS', meta={'current': processed_invoices, 'total': total_invoices})

    # Final message
    self.update_state(state='SUCCESS', meta={
        'current': total_invoices,
        'total': total_invoices,
        'message': "Textman invoice processing completed",
        'errors': errors  # Ensure errors are included in the meta response
    })

    # Step 1: Create a subquery to find the most recent record for each invoice_id
    latest_records_subquery = db.session.query(
        SupplierInvoiceRecord.invoice_id,
        func.max(SupplierInvoiceRecord.run_time).label("latest_run_time")
    ).filter(SupplierInvoiceRecord.invoice_id != None).group_by(SupplierInvoiceRecord.invoice_id).subquery()

    # Step 2: Create a subquery to find the IDs of older records
    subquery = db.session.query(
        SupplierInvoiceRecord.id
    ).join(
        latest_records_subquery,
        and_(
            SupplierInvoiceRecord.invoice_id == latest_records_subquery.c.invoice_id,
            SupplierInvoiceRecord.run_time < latest_records_subquery.c.latest_run_time
        )
    ).subquery()

    # Step 3: Use subquery.select() explicitly in the `.in_()` filter
    db.session.query(SupplierInvoiceRecord).filter(
        SupplierInvoiceRecord.id.in_(subquery.select())
    ).delete(synchronize_session=False)

    # Commit the changes
    db.session.commit()
    
    # Prepare final message
    if errors:
        return {
            "message": f"Textman invoice processing completed with {len(errors)} errors.",
            "errors": errors
        }
    else:
        return {"message": "Textman invoice processing completed successfully for all tenants."}



@shared_task(bind=True)
def process_eden_farm_task(self, user_id):
    # Fetch the user by user_id
    user = User.query.get(user_id)

    # Fetch all tenants associated with this user from DomPurchaseInvoicesTenant
    tenants = DomPurchaseInvoicesTenant.query.filter_by(user_id=user_id).all()

    # Extract tenant names to pass into the invoice fetching function
    tenant_names = [tenant.tenant_name for tenant in tenants]

    # Fetch Eden Farm invoices and credit notes for these tenants
    eden_farm_invoices= get_invoices_and_credit_notes(user, tenant_names, "Eden Farm")


    # Calculate total invoices to process
    total_invoices = len(eden_farm_invoices)
    processed_invoices = 0

    # Separate invoices and credit notes
    invoices_to_process = [invoice for invoice in eden_farm_invoices if invoice['xero_type'] == "invoice"]
    credit_notes_to_process = [invoice for invoice in eden_farm_invoices if invoice['xero_type'] == "credit_note"]

    # Track errors
    errors = []

    # Process the invoices for each tenant
    for invoice in invoices_to_process:
        data = extract_eden_farm_invoice_data(user, invoice)
        
        # Check if there are any errors
        invoice_errors = data.get("errors", [])
        if invoice_errors:
            errors.extend(invoice_errors)

            new_record = SupplierInvoiceRecord(
                user_id=user.id,
                store_name=invoice["tenant_name"],
                invoice_type="Eden Farm",
                invoice_number=invoice["invoice_reference"],
                invoice_id=invoice['invoice_id'],
                errors=', '.join(invoice_errors),
                triggered_by="manual",
                date_of_invoice=invoice["invoice_date"]
            )
            db.session.add(new_record)
            db.session.commit()

            add_log(f"Error processing Eden Farm invoice {invoice['invoice_id']}: {', '.join(invoice_errors)}", log_type="error", user_id=user.id)
            continue  # Skip this invoice due to errors

        # Process based on invoice type
        if data["invoice_type"] == "credit_memo":
            # Call the function to convert the credit memo to an invoice
            error = convert_invoice_to_credit_memo(data, user)
        else:
            # Call the function to assign tracking code to the invoice
            error = assign_tracking_code_to_invoice(data, user)

        new_record = SupplierInvoiceRecord(
            user_id=user.id,
            store_name=invoice["tenant_name"],
            invoice_type="Eden Farm",
            invoice_number=invoice["invoice_reference"],
            invoice_id=invoice['invoice_id'],
            triggered_by="manual",
            date_of_invoice=invoice["invoice_date"]
        )
        db.session.add(new_record)
        db.session.commit()

        # Update progress after processing each invoice
        processed_invoices += 1
        self.update_state(state='PROGRESS', meta={'current': processed_invoices, 'total': total_invoices})
    
    # Process the credit notes for each tenant
    for credit_note in credit_notes_to_process:
        data = extract_eden_farm_invoice_data(user, credit_note)

        # Check if there are any errors
        credit_note_errors = data.get("errors", [])
        if credit_note_errors:
            errors.extend(credit_note_errors)

            new_record = SupplierInvoiceRecord(
                user_id=user.id,
                store_name=invoice["tenant_name"],
                invoice_type="Eden Farm",
                invoice_number=credit_note["invoice_reference"],
                invoice_id=credit_note['invoice_id'],
                errors=', '.join(credit_note_errors),
                triggered_by="manual",
                date_of_invoice=credit_note["invoice_date"]
            )
            db.session.add(new_record)
            db.session.commit()

            add_log(f"Error processing Eden Farm credit note {credit_note['invoice_id']}: {', '.join(credit_note_errors)}", log_type="error", user_id=user.id)
            continue  # Skip this credit note due to errors

        # Process based on credit note type
        if data["invoice_type"] == "invoice":
            # Call the function to convert the credit memo to an invoice
            error = convert_credit_memo_to_invoice(data, user)
        else:
            # Call the function to assign tracking code to the credit memo
            error = assign_tracking_code_to_credit_note(data, user)
            add_log(f"Eden Farm credit note {credit_note['invoice_id']} is a valid credit memo.", log_type="general", user_id=user.id)


        new_record = SupplierInvoiceRecord(
            user_id=user.id,
            store_name=credit_note["tenant_name"],
            invoice_type="Eden Farm",
            invoice_number=credit_note["invoice_reference"],
            invoice_id=credit_note['invoice_id'],
            triggered_by="manual",
            date_of_invoice=credit_note["invoice_date"]
        )
        db.session.add(new_record)
        db.session.commit()

        # Update progress after processing each credit note
        processed_invoices += 1
        self.update_state(state='PROGRESS', meta={'current': processed_invoices, 'total': total_invoices})

    # Final message and update state
    self.update_state(state='SUCCESS', meta={
        'current': total_invoices,
        'total': total_invoices,
        'message': "Eden Farm invoice processing completed",
        'errors': errors  # Ensure errors are included in the meta response
    })

    # Step 1: Create a subquery to find the most recent record for each invoice_id
    latest_records_subquery = db.session.query(
        SupplierInvoiceRecord.invoice_id,
        func.max(SupplierInvoiceRecord.run_time).label("latest_run_time")
    ).filter(SupplierInvoiceRecord.invoice_id != None).group_by(SupplierInvoiceRecord.invoice_id).subquery()

    # Step 2: Create a subquery to find the IDs of older records
    subquery = db.session.query(
        SupplierInvoiceRecord.id
    ).join(
        latest_records_subquery,
        and_(
            SupplierInvoiceRecord.invoice_id == latest_records_subquery.c.invoice_id,
            SupplierInvoiceRecord.run_time < latest_records_subquery.c.latest_run_time
        )
    ).subquery()

    # Step 3: Use subquery.select() explicitly in the `.in_()` filter
    db.session.query(SupplierInvoiceRecord).filter(
        SupplierInvoiceRecord.id.in_(subquery.select())
    ).delete(synchronize_session=False)

    # Commit the changes
    db.session.commit()
    
    # Prepare the final response message
    if errors:
        return {
            "message": f"Eden Farm invoice processing completed with {len(errors)} errors.",
            "errors": errors
        }
    else:
        return {"message": "Eden Farm invoice processing completed successfully for all tenants."}



# Celery task to upload invoices to Xero
@shared_task(bind=True)
def upload_recharge_invoices_xero_task(self, user_id, purchase_csv_content, breakdown_csv_content, sales_invoices_csv_content):
    try:
        # Load user from the database
        user = User.query.get(user_id)

        if not user:
            raise ValueError(f"User with id {user_id} not found.")
        
        # Retrieve the stored CSV content from the session
        purchase_csv_content = purchase_csv_content
        breakdown_csv_content = breakdown_csv_content
        sales_invoices_csv_content = sales_invoices_csv_content

        if not purchase_csv_content or not breakdown_csv_content:
            raise ValueError("CSV content is missing in session")

        # Process the purchase CSV into a DataFrame
        purchase_csv_io = io.StringIO(purchase_csv_content)
        purchase_df = pd.read_csv(purchase_csv_io)
        
        sales_csv_io = io.StringIO(sales_invoices_csv_content)
        sales_invoice_df  = pd.read_csv(sales_csv_io)

        # Check if the 'DueDate' and 'InvoiceNumber' columns exist and log errors if missing
        due_date = None
        invoice_number = None

        # Check if the 'DueDate' and 'InvoiceNumber' columns exist and log errors if missing
        due_date = purchase_df.get('DueDate', pd.Series([None])).iloc[0]
        invoice_number = purchase_df.get('InvoiceNumber', pd.Series([None])).iloc[0]

        if pd.isna(due_date):
            raise ValueError("DueDate is missing in the 'purchase_csv'.")
        if pd.isna(invoice_number):
            raise ValueError("InvoiceNumber is missing in the 'purchase_csv'.")
        
         # Process the breakdown CSV into a DataFrame
        breakdown_df = pd.read_csv(io.StringIO(breakdown_csv_content))
    

        invoices_data = {}
        sales_invoices_data = {}
        sales_invoice_numbers = {}
        invoice_numbers = {}
        company_codes = {}

        # Query all valid companies from TrackingCategoryModel
        valid_companies = db.session.query(TrackingCategoryModel.tenant_name).distinct().all()
        valid_companies = {company[0] for company in valid_companies}  # Set of valid company names

        # Lists to store the result
        successful_companies = []
        failed_companies = []
        already_processed_companies = []
        sales_successful_companies = []
        sales_failed_companies = []
        sales_already_processed_companies = []


        # Iterate over the DataFrame rows to create line items for Xero
        for _, row in purchase_df.iterrows():
            company_name = row.get('Company Name')
            invoice_number = row.get('InvoiceNumber')  # Extract the InvoiceNumber for each respective company row

            if company_name not in valid_companies:
                current_app.logger.error(f"Company '{company_name}' not found in tracking categories. Skipping row.")
                failed_companies.append(company_name)
                continue

            description = row.get('Description')
            quantity = float(row.get('Quantity', 1))  # Default to 1 if not provided
            unit_amount = float(row.get('UnitAmount'))
            account_code = row.get('AccountCode')
            tracking_option_name = row.get('TrackingOption1')

            # Query the TrackingCategoryModel to get the tracking_category_id and tracking_option_id
            tracking_record = TrackingCategoryModel.query.filter_by(tracking_category_option=tracking_option_name).first()
            if not tracking_record:
                current_app.logger.error(f"Tracking option '{tracking_option_name}' not found for company '{company_name}'.")
                failed_companies.append(company_name)
                continue

            if company_name not in invoices_data:
                invoices_data[company_name] = []
                invoice_numbers[company_name] = invoice_number

            line_item_tracking = LineItemTracking(
                tracking_category_id=tracking_record.tracking_category_id,
                tracking_option_id=tracking_record.tracking_option_id
            )

            invoices_data[company_name].append(
                LineItem(
                    description=description,
                    quantity=quantity,
                    unit_amount=unit_amount,
                    account_code=account_code,
                    tracking=[line_item_tracking]
                )
            )

        # Iterate over the sales innvoie DataFrame rows to create line items for Xero
        for _, row in sales_invoice_df.iterrows():
            company_name = row.get('ContactName')
            invoice_number = row.get('InvoiceNumber')  # Extract the InvoiceNumber for each respective company row
            description = row.get('Description')
            quantity = float(row.get('Quantity', 1))  # Default to 1 if not provided
            unit_amount = float(row.get('UnitAmount'))
            account_code = row.get('AccountCode')


            # Query the TrackingCategoryModel to get the tracking_category_id and tracking_option_id
            tracking_record = TrackingCategoryModel.query.filter_by(tracking_category_option=tracking_option_name).first()
            if not tracking_record:
                current_app.logger.error(f"Tracking option '{tracking_option_name}' not found for company '{company_name}'.")
                failed_companies.append(company_name)
                continue

            if company_name not in sales_invoices_data:
                sales_invoices_data[company_name] = []
                sales_invoice_numbers[company_name] = invoice_number

            line_item_tracking = LineItemTracking(
                tracking_category_id=tracking_record.tracking_category_id,
                tracking_option_id=tracking_record.tracking_option_id
            )

            sales_invoices_data[company_name].append(
                LineItem(
                    description=description,
                    quantity=quantity,
                    unit_amount=unit_amount,
                    account_code=account_code,
                    tax_type="OUTPUT2"
                )
            )
        

        contact_data = get_all_contacts(user)
        tenant_info = {}
        sales_tenant_info = {}


        for company in invoices_data.keys():
            company_stripped = company.strip().lower()
            tenant_contact_info = next((tenant for tenant in contact_data if tenant['tenant_name'].lower().strip() == company_stripped), None)

            if not tenant_contact_info:
                failed_companies.append(company)
                continue

            tenant_id = tenant_contact_info['tenant_id']
            contact_record = next((contact for contact in tenant_contact_info['contacts'] if user.username in contact['contact_name']), None)

            if not contact_record:
                failed_companies.append(company)
                continue

            tenant_info[company] = {
                'tenant_id': tenant_id,
                'contact_id': contact_record['contact_id']
            }

        # Second loop to match tenant names with user.company_name and store tenant ID, contact ID, and company name using the company as the key
        for company in sales_invoices_data.keys():
            company_stripped = company.strip().lower()
            # Find tenant matching the user's company name
            tenant_contact_info = next((tenant for tenant in contact_data if tenant['tenant_name'].lower().strip() == user.company_name.lower().strip()), None)

            if not tenant_contact_info:
                add_log(f"Missing [{company} company contact in management company]", log_type="error", user_id=user.id)
                raise ValueError(f"Missing [{company} company contact in management company]")
                

            # Check if any contact names in the tenant match the company name in sales_invoices_data.keys()
            contact_record = next((contact for contact in tenant_contact_info['contacts'] if contact['contact_name'].lower().strip() == company_stripped), None)

            if not contact_record:
                add_log(f"Missing [{company} company contact in management company]", log_type="error", user_id=user.id)
                raise ValueError(f"Missing [{company} company contact in management company]")
                

            # Store the data in sales_tenant_info with the company as the key
            sales_tenant_info[company] = {
                'tenant_id': tenant_contact_info['tenant_id'],
                'contact_id': contact_record['contact_id'],
                'company_name': company
            }

        # Print or use sales_tenant_info as needed
        print("Sales Tenant Info:", sales_tenant_info)

        
        
        

        # Check if all companies have valid company codes
        for company in invoices_data.keys():
            # Get the company code from the Company table for the current company
            company_record = Company.query.filter_by(company_name=company, user_id=user_id).first()
            if not company_record or not company_record.company_code:
                current_app.logger.error(f"Company record or company code not found for '{company}'.")
                raise ValueError(f"Missing company code for company [{company}]")

            # Store the company code for later use in the invoice processing
            company_codes[company] = company_record.company_code



        for company, line_items in invoices_data.items():
            invoice_number = invoice_numbers[company]
            file_name = (company + " Breakdown " + "(" + invoice_number + ")" + ".csv" )
            tenant_id = tenant_info[company]['tenant_id']
            contact_id = tenant_info[company]['contact_id']


            # Retrieve the pre-validated company code
            company_code = company_codes.get(company)

            # Filter the breakdown DataFrame for the current company code
            filtered_breakdown_df = breakdown_df[breakdown_df['Company Code'] == company_code]

            # Sort the filtered breakdown DataFrame by 'Account Code Per Business'
            filtered_breakdown_df = filtered_breakdown_df.sort_values(by='Account Code Per Business')

            # Calculate the net total per account code per business
            net_total_summary = filtered_breakdown_df.groupby(['Account Code Per Business'])['Net'].sum().reset_index()

            # Create the summary table with two columns: Account Code and Net Total
            summary_table = pd.DataFrame({
                'Account Code Per Business': net_total_summary['Account Code Per Business'],
                'Net Total': net_total_summary['Net']
            })

            # Insert a blank row for spacing
            blank_row = pd.DataFrame([[''] * len(filtered_breakdown_df.columns)], columns=filtered_breakdown_df.columns)

            # Append blank row after the filtered breakdown data
            summary_df = pd.concat([filtered_breakdown_df, blank_row], ignore_index=True)

            # Convert the filtered breakdown DataFrame to CSV
            breakdown_csv_content = summary_df.to_csv(index=False)

            # Convert the summary table to CSV
            summary_csv_content = summary_table.to_csv(index=False)

            # Combine the breakdown CSV content and the summary table
            csv_content = breakdown_csv_content + '\n\n' + summary_csv_content

            # Convert the combined content to a byte array
            byte_array_csv_content = csv_content.encode('utf-8')


            try:
                result = post_recharge_purchase_invoice_xero(
                    tenant_id=tenant_id,
                    line_items=line_items,
                    contact_id=contact_id,
                    file_name=file_name,
                    file_content=byte_array_csv_content,
                    user=user,
                    end_date=due_date,
                    invoice_number=invoice_number
                )

                if result == "ALREADY CREATED":
                    already_processed_companies.append(company)
                else:
                    successful_companies.append(company)

            except Exception as e:
                current_app.logger.error(f"Error creating invoice for company '{company}': {str(e)}")
                failed_companies.append(company)

        
        for company, line_items in sales_invoices_data.items():
            try:
                # Retrieve invoice number and tenant/contact info
                invoice_number = invoice_numbers[company]
                file_name = f"{company} Breakdown ({invoice_number}).csv"
                tenant_id = sales_tenant_info[company]['tenant_id']
                contact_id = sales_tenant_info[company]['contact_id']

                # Retrieve the pre-validated company code
                company_code = company_codes.get(company)

                # Filter the breakdown DataFrame for the current company code
                filtered_breakdown_df = breakdown_df[breakdown_df['Company Code'] == company_code]

                # Sort the filtered breakdown DataFrame by 'Account Code Per Business'
                filtered_breakdown_df = filtered_breakdown_df.sort_values(by='Account Code Per Business')

                # Calculate the net total per account code per business
                net_total_summary = filtered_breakdown_df.groupby(['Account Code Per Business'])['Net'].sum().reset_index()

                # Create the summary table with two columns: Account Code and Net Total
                summary_table = pd.DataFrame({
                    'Account Code Per Business': net_total_summary['Account Code Per Business'],
                    'Net Total': net_total_summary['Net']
                })

                # Insert a blank row for spacing
                blank_row = pd.DataFrame([[''] * len(filtered_breakdown_df.columns)], columns=filtered_breakdown_df.columns)

                # Append blank row after the filtered breakdown data
                summary_df = pd.concat([filtered_breakdown_df, blank_row], ignore_index=True)

                # Convert the filtered breakdown DataFrame to CSV
                breakdown_csv_content = summary_df.to_csv(index=False)

                # Convert the summary table to CSV
                summary_csv_content = summary_table.to_csv(index=False)

                # Combine the breakdown CSV content and the summary table
                csv_content = breakdown_csv_content + '\n\n' + summary_csv_content

                # Convert the combined content to a byte array
                byte_array_csv_content = csv_content.encode('utf-8')

                # Post the sales invoice to Xero
                result = post_recharge_sales_invoice_xero(
                    tenant_id=tenant_id,
                    line_items=line_items,
                    contact_id=contact_id,
                    file_name=file_name,
                    file_content=byte_array_csv_content,
                    user=user,
                    end_date=due_date,
                    invoice_number=invoice_number
                )

                # Track result
                if result == "ALREADY CREATED":
                    sales_already_processed_companies.append(company)
                else:
                    sales_successful_companies.append(company)

            except Exception as e:
                current_app.logger.error(f"Error creating invoice for company '{company}': {str(e)}")
                sales_failed_companies.append(company)


        # Step 4: Remove duplicates from both lists
        successful_companies = list(set(successful_companies))
        failed_companies = list(set(failed_companies))
        already_processed_companies = list(set(already_processed_companies))


        # Step 4: Log the successful and failed companies
        log_message = f"Successful companies: {', '.join(successful_companies)}; Failed companies: {', '.join(failed_companies)}"
        add_log(log_message, log_type="recharge_invoices", user_id=user.id)

        return {
            "status": "success",
            "message": "Invoices uploaded to Xero successfully.",
            "successful_companies": successful_companies,
            "failed_companies": failed_companies,
            "already_processed_companies": already_processed_companies,
            "sales_successful_companies": sales_successful_companies,
            "sales_failed_companies": sales_failed_companies,
            "sales_already_processed_companies": sales_already_processed_companies

        }

    except Exception as e:
        current_app.logger.error(f"Error in upload_recharge_invoices_xero_task: {str(e)}")
        self.update_state(state='FAILURE', meta={'error': str(e), 'exc_type': type(e).__name__})
        raise



##########################Stock######################################

@shared_task(bind=True)
def process_inventory_task(self, user_id):
    """
    Processes the uploaded Excel files to update the inventory records and produce journals in Xero.
    """
    from datetime import datetime

    user = User.query.get(user_id)
    task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

    if task_status:
        task_status.status = 'in_progress'
        task_status.result = "Task started."
        db.session.commit()

    data, management_tenant_id = get_inbox_files_from_management_company(user)
    error_messages = []
    to_be_processed = []

    if data:
        for tenant in data:
            tenant_id = tenant['tenant_id']

            # Create folders if they don't exist
            processed_folder = create_folder_if_not_exists('Stock - Processed', tenant_id, user)
            rejected_folder = create_folder_if_not_exists('Stock - Rejected', tenant_id, user)

            processed_folder_id = processed_folder.id
            rejected_folder_id = rejected_folder.id

            for file in tenant['files']:
                if "Stock_Record" in file['file_name']:
                    to_be_processed.append((file, tenant_id, processed_folder_id, rejected_folder_id))

        # Process each file
        for file, tenant_id, processed_folder_id, rejected_folder_id in to_be_processed:
            file_content = fetch_file_content(user, tenant_id, file['file_id'])
            result = process_inventory_file(
                user=user,
                file=file,
                processed_folder_id=processed_folder_id,
                rejected_folder_id=rejected_folder_id,
                tenant_id=tenant_id,
                file_content=file_content
            )
            if "error" in result:
                error_messages.append(result["error"])

    if task_status:
        task_status.status = 'completed' if not error_messages else 'failed'
        task_status.result = "Task completed with errors." if error_messages else "Task completed successfully."
        task_status.progress = 100
        db.session.commit()

    return {"errors": error_messages} if error_messages else {"message": "Task completed successfully."}

@shared_task(bind=True)
def create_inventory_journals(self, current_month_records, previous_month_records, user_id, month, year):
    from app.models import TrackingCategoryModel
    import pandas as pd

    user = User.query.get(user_id)

    # Step 1: Create DataFrames from current and previous month records
    current_df = pd.DataFrame(current_month_records)
    previous_df = pd.DataFrame(previous_month_records)

    # Step 2: Merge DataFrames on 'store_name'
    merged_df = pd.merge(
        current_df, 
        previous_df, 
        on='store_name', 
        how='left', 
        suffixes=('_current', '_previous')
    )

    # Step 3: Calculate the difference column
    merged_df['difference'] = merged_df['amount_current'] - merged_df['amount_previous']

    # Step 4: Query tracking codes from the database
    tracking_codes = TrackingCategoryModel.query.filter_by(user_id=user_id).all()
    tracking_code_map = {
        tc.tracking_category_option: {
            "company_name": tc.tenant_name,
            "tracking_option_id": tc.tracking_option_id,
            "tracking_category_id": tc.tracking_category_id
        }
        for tc in tracking_codes
    }

    # Step 5: Map company_name, tracking_option_id, and tracking_category_id to merged DataFrame
    merged_df['company_name'] = merged_df['store_name'].map(
        lambda store_name: tracking_code_map.get(store_name, {}).get('company_name', None)
    )
    merged_df['tracking_option_id'] = merged_df['store_name'].map(
        lambda store_name: tracking_code_map.get(store_name, {}).get('tracking_option_id', None)
    )
    merged_df['tracking_category_id'] = merged_df['store_name'].map(
        lambda store_name: tracking_code_map.get(store_name, {}).get('tracking_category_id', None)
    )

    # Step 6: Finalize the columns
    final_df = merged_df[[
        'store_name', 
        'amount_current', 
        'amount_previous', 
        'difference', 
        'company_name', 
        'tracking_option_id',
        'tracking_category_id', 
    ]]

    # Step 7: Pass the final DataFrame to the processing function
    create_inventory_journals_in_xero(final_df, user, month, year)

    return {"status": "success"}








##########################RECHARGING######################################

import os

from io import BytesIO
import zipfile
import pandas as pd
from app.models import User, AccountTransaction  # Import your models
from datetime import datetime
import calendar
from flask import Blueprint, render_template, session, jsonify, request, flash, redirect, url_for, send_file, Response
from openpyxl import load_workbook
from sqlalchemy.orm import joinedload


@shared_task(bind=True)
def process_recharging_task(self, user_id, selected_month, selected_year, last_invoice_number):
    try:
        # Fetch the TaskStatus record for the current task
        task_status = TaskStatus.query.filter_by(task_id=self.request.id).first()

        if task_status:
            task_status.status = 'in_progress'
            task_status.result = "Task started."
            db.session.commit()

        user = User.query.get(user_id)
        current_company = user.company_name
        current_username = user.username

        # Log when the process starts
        add_log("Processing Macro button pressed", log_type="general", user_id=user_id)

       
        lastInvoiceNumber = last_invoice_number 
        #lastInvoiceNumber = manual set

        selected_month_str = f"{selected_year}-{selected_month:02d}"
        selected_month_year_str_filenames = f"{datetime(1900, selected_month, 1).strftime('%B')} {selected_year}"

        add_log(f"Selected month: {selected_month_str}", log_type="general", user_id=user_id)

        # Query the database for account transactions for the logged-in user
        transactions = AccountTransaction.query.filter_by(user_id=user_id).all()
        

        # Example of handling an error case
        if not transactions:
            add_log("No data found for the current user.", log_type="error", user_id=user_id)

            if task_status:
                task_status.status = 'failed'
                task_status.result = "error: No data found for the current month and year selected. Please check the logs."
                db.session.commit()

            return {"status": "error", "message":"No data found for the current user. Please check the logs."}

        add_log(f"Found {len(transactions)} transactions for the user.", log_type="general", user_id=user_id)

        # Convert the transaction records to a pandas DataFrame
        transaction_data = [
            {
                'Date': t.date,
                'Source': t.source,
                'Contact': t.contact,
                'Description': t.description,
                'Reference': t.reference,
                'Debit': t.debit,
                'Credit': t.credit,
                'Gross': t.gross,
                'Net': t.net,
                'VAT': t.vat,
                'Account Code': t.account_code,
                'Account': t.account,
                'tracking_group1': t.tracking_group1,
                'tracking_group2': t.tracking_group2
            } for t in transactions
        ]
        df_unfiltered = pd.DataFrame(transaction_data)

        # Calculate the sum of the 'Net' column
        net_sum = df_unfiltered['Net'].sum()

        # Print the sum of the 'Net' column
        #print(f"Sum of 'Net' column: {net_sum}")

        if df_unfiltered.empty:
            add_log("No data found in DataFrame after converting transactions.", log_type="error", user_id=user_id)

            if task_status:
                task_status.status = 'failed'
                task_status.result = "error: No data found in data.csv."
                db.session.commit()

            return {"status": "error", "message": "No data found in data.csv."}

        # Filter by the selected month and year

        # Step 1: Try to parse the 'Date' with the first format ('%Y-%m-%d')
        df_unfiltered['Date'] = pd.to_datetime(df_unfiltered['Date'], format='%Y-%m-%d', errors='coerce')

        # Step 2: For rows where the 'Date' is NaT (failed to parse), try the second format ('%d %b %Y')
        df_unfiltered['Date'] = df_unfiltered['Date'].fillna(pd.to_datetime(df_unfiltered['Date'], format='%d %b %Y', errors='coerce'))
        

        df  = df_unfiltered[(df_unfiltered['Date'].dt.month == selected_month) &
                                    (df_unfiltered['Date'].dt.year == selected_year)]


        if df.empty:
            add_log(f"No transactions found for {selected_month_str}.", log_type="error", user_id=user_id)

            if task_status:
                task_status.status = 'failed'
                task_status.result = "error: No data found for the current month and year selected. Please check the logs."
                db.session.commit()

            return {"status": "error", "message":"No data found for the current user. Please check the logs."}

        else:
            add_log(f"Found {len(df)} transactions for {selected_month_str}.", log_type="general", user_id=user_id)


        # Further processing with logs
        modified_df = combine_last_two_columns(df)
        columns_to_drop = ['Debit', 'Credit', 'Gross', 'VAT']
        modified_df = modified_df.drop(columns=columns_to_drop, errors='ignore')
        modified_df['Account Code Per Business'] = ''
        modified_df['Account Code Per Business Description'] = ''
        modified_df['Net'] = modified_df['Net'].astype(float)

        add_log("DataFrame modified successfully. Dropped unnecessary columns and added business account codes.", log_type="general", user_id=user_id)

        # Get all account codes per DMS and map them to business account codes
        account_codes_per_dms = (
            db.session.query(
                AccountCodesPerDMS.account_code_per_dms,
                AccountCodesPerBusiness.account_code_per_business,
                AccountCodesPerDMS.descriptor_per_dms,
                AccountCodesPerBusiness.descriptor_per_business
            )
            .join(AccountCodesPerBusiness, AccountCodesPerDMS.business_id == AccountCodesPerBusiness.id)
            .filter(AccountCodesPerDMS.user_id == user_id)
            .all()
        )

        add_log(f"Fetched {len(account_codes_per_dms)} account codes per DMS.", log_type="general", user_id=user_id)

        # Query the database for companies associated with the user
        companies = Company.query.filter_by(user_id=user_id).all()

        # Extract company names and codes into a list of dictionaries
        data = [
            {
                'company_name': company.company_name,
                'company_code': company.company_code
            }
            for company in companies
        ]

        # Create a pandas DataFrame from the data
        companies_df = pd.DataFrame(data)

        add_log(f"Fetched {len(companies_df)} companies.", log_type="general", user_id=user_id)

        # Convert the query result to a DataFrame
        account_codes_per_dms_data = [
            {
                'Account Code Per DMS': code_per_dms[0],
                'Account Code Per Business': code_per_dms[1],
                'Descriptor Per DMS': code_per_dms[2],
                'Descriptor Per Business': code_per_dms[3]
            } for code_per_dms in account_codes_per_dms
        ]
        account_codes_per_dms_df = pd.DataFrame(account_codes_per_dms_data)

        # Ensure columns are integers
        account_codes_per_dms_df['Account Code Per DMS'] = account_codes_per_dms_df['Account Code Per DMS'].astype(int)
        account_codes_per_dms_df['Account Code Per Business'] = account_codes_per_dms_df['Account Code Per Business'].astype(int)

        # Further processing (Tracking Codes, Group Codes, etc.)
        add_log("Tracking codes and group codes processing started.", log_type="general", user_id=user_id)

        tracking_codes = TrackingCode.query.filter_by(user_id=user_id).with_entities(TrackingCode.tracking_code).all()
        tracking_codes = [code[0] for code in tracking_codes if code]

        group_tracking_codes = GroupTrackingCode.query.filter_by(user_id=user_id).with_entities(GroupTrackingCode.group_code).all()
        group_tracking_codes = [code[0] for code in group_tracking_codes]

        group_tracking_code_mappings = GroupTrackingCode.query.filter_by(user_id=user_id).options(joinedload(GroupTrackingCode.tracking_codes)).all()

        group_tracking_code_map = {}
        for group in group_tracking_code_mappings:
            group_code = group.group_code
            tracking_codes_associated = [tc.tracking_code for tc in group.tracking_codes]
            group_tracking_code_map[group_code] = tracking_codes_associated


        # Process each row of modified_df
        processed_rows = []
        for _, row in modified_df.iterrows():
            tracking_code = row['Tracking Code']
            net_value = row['Net']
            if tracking_code in tracking_codes:
                company_code = tracking_code.split(' -')[0]
                row['Company Code'] = company_code
                processed_rows.append(row)
            elif tracking_code in group_tracking_codes:
                assigned_codes = group_tracking_code_map[tracking_code]
                num_assigned_codes = len(assigned_codes)
                divided_net_value = round(net_value / num_assigned_codes, 2)
                total_rounded_value = divided_net_value * (num_assigned_codes - 1)
                remainder = round(net_value - total_rounded_value, 2)
                for i, code in enumerate(assigned_codes):
                    new_row = row.copy()
                    if i == num_assigned_codes - 1:
                        new_row['Net'] = remainder
                    else:
                        new_row['Net'] = divided_net_value
                    new_row['Tracking Code'] = code
                    company_code = code.split(' -')[0]
                    new_row['Company Code'] = company_code
                    new_row['Description'] = f"{new_row['Description']} - Total Invoice Amount = {net_value}, Tracking Code = {tracking_code}"
                    processed_rows.append(new_row)
            else:
                add_log(f"Error: Tracking code {tracking_code} not found in any tracking codes.", log_type="error", user_id=user_id)
                
                if task_status:
                    task_status.status = 'failed'
                    task_status.result = f"error: Tracking code {tracking_code} not found in any tracking codes."
                    db.session.commit()

                return {"status": "error", "message": f"Tracking code {tracking_code} not found in any tracking codes."}

        add_log(f"Processed {len(processed_rows)} rows of tracking data.", log_type="general", user_id=user_id)

        final_df = pd.DataFrame(processed_rows)

        # Match account codes with business codes
        for index, row in final_df.iterrows():
            account_code = row['Account Code']
            matching_row = account_codes_per_dms_df[account_codes_per_dms_df['Account Code Per DMS'] == account_code]
            if matching_row.empty:
                add_log(f"No matching account code per business found for {account_code}.", log_type="error", user_id=user_id)

                if task_status:
                    task_status.status = 'failed'
                    task_status.result = f"error: No matching account code per business found for {account_code}"
                    db.session.commit()

                return {"status": "error", "message":f"No matching account code per business found for {account_code}"}
            else:
                account_code_per_business = matching_row.iloc[0]['Account Code Per Business']
                descriptor_per_business = matching_row.iloc[0]['Descriptor Per Business']
                final_df.at[index, 'Account Code Per Business'] = account_code_per_business
                final_df.at[index, 'Account Code Per Business Description'] = descriptor_per_business

        add_log("Account codes matched successfully.", log_type="general", user_id=user_id)

        # Generate final files, zip them, and return to the user
        add_log("Generating invoices and zipping them.", log_type="general", user_id=user_id)

        

        # Create a new DataFrame for the sales invoice
        sales_invoice_df = pd.DataFrame(columns=['ContactName','EmailAdress','POAddressLine1','POAddressLine2','POAddressLine3','POAddressLine4','POCity','PORegion','POPostalCode','POCountry', 'InvoiceNumber', 'Company Code', 'InvoiceDate', 'DueDate','Total','InventoryItemCode', 'Description', 'Quantiy', 'UnitAmount','Discount','AccountCode', 'TaxType'])
        
        starting_invoice_number = lastInvoiceNumber
        row_count = 0

        # Get the current date
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_month = datetime.now().strftime("%B")
        current_year = datetime.now().strftime("%Y")

        # Calculate the last day of the selected month and year
        last_day_of_month = calendar.monthrange(selected_year, selected_month)[1]

        # Create the invoice date as the last day of the selected month and year
        invoice_date = datetime(selected_year, selected_month, last_day_of_month).strftime("%d/%m/%Y")

        selected_month_name = calendar.month_name[selected_month]


        # Process each row in the modified DataFrame
        for _, row in final_df.iterrows():
            company_code = row['Company Code']
            net_amount = row['Net']

            # Find the company name from companies.csv
            # Get the company name where the company_code matches
            matching_company = companies_df.loc[companies_df['company_code'] == company_code, 'company_name']

            # Check if there are any matches
            if not matching_company.empty:
                company_name = matching_company.values[0]
            else:
                company_name = None  # Set company_name to None if no match is found
                add_log(f"Error: No company found for company code '{company_code}'", log_type="error", user_id=user_id)


            # Check if the company name already exists in the invoice DataFrame
            if not sales_invoice_df[(sales_invoice_df['ContactName'] == company_name) & (sales_invoice_df['Company Code'] == company_code)].empty:
                # Add the net amount to the existing net amount
                sales_invoice_df.loc[(sales_invoice_df['ContactName'] == company_name) & (sales_invoice_df['Company Code'] == company_code), 'UnitAmount'] +=net_amount
            else:
                #  Add a new row
                new_row = pd.DataFrame({
                    'ContactName': [company_name],
                    'InvoiceNumber': [str(current_username)+' - ' + str(starting_invoice_number + row_count)],
                    'Quantity': ['1'],
                    'TaxType':['20% (VAT on Income)'],
                    'Description': ['HO and Management cost recharges for the month of ' + str(selected_month_name) + ' ' + str(selected_year)],
                    'Company Code': [company_code],
                    'InvoiceDate': [invoice_date],
                    'DueDate': [invoice_date],
                    'UnitAmount': [net_amount],
                    'AccountCode': ['4000']
                })
                row_count = row_count + 1
                sales_invoice_df = pd.concat([sales_invoice_df, new_row], ignore_index=True)
        



        # Ensure all values in the 'Net' column are rounded to two decimal places
        sales_invoice_df['UnitAmount'] = sales_invoice_df['UnitAmount'].round(2)


        #st.write("Sales Invoice Data:")
        #st.dataframe(sales_invoice_df)

        # Convert sales_invoice_df to CSV
        csv_sales = sales_invoice_df.to_csv(index=False).encode('utf-8')

        
    # Button to produce purchase invoices
    #if st.button("Produce Purchase Invoices"):


        # Create a new DataFrame for the purchase invoice
        purchase_invoice_df = pd.DataFrame(columns=['ContactName','EmailAddress','POAddressLine1','POAddressLine2','POAddressLine3','POAddressLine4','POCity','PORegion','POPostalCode','POCountry','InvoiceNumber', 'InvoiceDate', 'DueDate', 'Total', 'InventoryItemCode', 'Description', 'Quantity', 'UnitAmount', 'AccountCode', 'TaxType', 'TaxAmount', 'TrackingName1', 'TrackingOption1', 'TrackingName2', 'TrackingOption2', 'Currency', 'Company Code', 'Company Name'])

        # Process each row in the final DataFrame
        for _, row in final_df.iterrows():
            company_code = row['Company Code']
            net_amount = row['Net']
            #tracking_code = row['Tracking Code'].split('-')[1].strip() 
            tracking_code = '-'.join(row['Tracking Code'].split('-')[1:]).strip()

            account_code = row['Account Code Per Business']  # Assuming 'Account Code' is a column in your DataFrame


            # Find the company name from companies.csv
            company_name = companies_df.loc[companies_df['company_code'] == company_code, 'company_name'].values[0]

            # Check if the company code and account code already exist in the purchase invoice DataFrame
            existing_row = purchase_invoice_df[(purchase_invoice_df['Company Code'] == company_code) & (purchase_invoice_df['AccountCode'] == account_code) & (purchase_invoice_df['TrackingOption1'] == tracking_code) ]
            if not existing_row.empty:
                # Add the net amount to the existing net amount
                purchase_invoice_df.loc[(purchase_invoice_df['Company Code'] == company_code) & (purchase_invoice_df['AccountCode'] == account_code)  & (purchase_invoice_df['TrackingOption1'] == tracking_code) , 'UnitAmount'] += net_amount
            else:
                # Find the row in the sales_invoice_df with the matching company name
                matching_row = sales_invoice_df[sales_invoice_df['ContactName'] == company_name]

                # Retrieve the invoice number from the matching row
                if not matching_row.empty:
                    invoice_number = matching_row.iloc[0]['InvoiceNumber']
                else:
                    invoice_number = None  # or handle the case where no match is found
                    
                
                # Add a new row
                new_row = pd.DataFrame({
                    'ContactName': [current_company],
                    'InvoiceDate': [invoice_date],
                    'DueDate': [invoice_date],
                    'Quantity': ['1'],
                    'TaxType':['20% (VAT on Expenses)'],
                    'TrackingName1': ['Store'],
                    'InvoiceNumber': [invoice_number],
                    'Description': ['HO and Management cost recharges for the month of ' + str(selected_month_name) + ' ' + str(selected_year)],
                    'TrackingOption1': [tracking_code],
                    'Company Name': [company_name],
                    'Company Code': [company_code],
                    'AccountCode': [account_code],
                    'UnitAmount': [net_amount]
                })
                purchase_invoice_df = pd.concat([purchase_invoice_df, new_row], ignore_index=True)
            

        # Ensure all values in the 'Net' column are rounded to two decimal places
        purchase_invoice_df['UnitAmount'] = purchase_invoice_df['UnitAmount'].round(2)


        #st.write("Purchase Invoice Data:")
        #st.dataframe(purchase_invoice_df)
        # Create directories if they don't exist
        # 1. Create purchase_invoices.zip in memory
        purchase_invoices_zip_stream = BytesIO()
        with zipfile.ZipFile(purchase_invoices_zip_stream, 'w', zipfile.ZIP_DEFLATED) as purchase_zip:
            for company in purchase_invoice_df['Company Name'].unique():
                company_df = purchase_invoice_df[purchase_invoice_df['Company Name'] == company]
                # Get the invoice number from the first row of the company data
                invoice_number = company_df.iloc[0]['InvoiceNumber']
                # Add the invoice number at the beginning of the filename
                company_filename = f"{invoice_number}-{company}-{selected_month_year_str_filenames}.csv"
                csv_content = company_df.to_csv(index=False)
                purchase_zip.writestr(company_filename, csv_content)
        purchase_invoices_zip_stream.seek(0)
        zip_data1 = purchase_invoices_zip_stream.read()

        # 2. Create breakdown_invoices.zip in memory
        breakdown_invoices_zip_stream = BytesIO()
        with zipfile.ZipFile(breakdown_invoices_zip_stream, 'w', zipfile.ZIP_DEFLATED) as breakdown_zip:
            for company_code in final_df['Company Code'].unique():
                # Filter the dataframe for the current company
                company_df = final_df[final_df['Company Code'] == company_code]
                
                # Get the company name where the company_code matches
                matching_company_name = companies_df.loc[companies_df['company_code'] == company_code, 'company_name'].values

                # Check if matching_company_name is not empty and extract the first value
                if len(matching_company_name) > 0:
                    company_name_str = matching_company_name[0]
                else:
                    company_name_str = "Unknown Company"

                # Sort the transactions by 'Account Code per Business'
                company_df_sorted = company_df.sort_values(by='Account Code Per Business')

                # Initialize an empty DataFrame for breakdown with totals
                breakdown_with_totals = pd.DataFrame()

                # Loop through each 'Account Code per Business' and group data
                for code, group in company_df_sorted.groupby('Account Code Per Business'):
                    
                    # Sort each group by 'Net' in descending order
                    group = group.sort_values(by='Net', ascending=False)
        
                    # Calculate the total for 'Net' column, rounded to 2 decimal places
                    total_net = round(group['Net'].sum(), 2)

                    # Add a 'Total Net' column, setting the total value only on the last row of the group
                    group['Total Net'] = [None] * (len(group) - 1) + [total_net]

                    # Concatenate the group with breakdown_with_totals
                    breakdown_with_totals = pd.concat([breakdown_with_totals, group], ignore_index=True)

                    # Add an empty row for separation between different account codes
                    empty_row = pd.Series([None] * len(breakdown_with_totals.columns), index=breakdown_with_totals.columns)
                    breakdown_with_totals = pd.concat([breakdown_with_totals, pd.DataFrame([empty_row])], ignore_index=True)

                # Check and print the column names for debugging
                #print("Columns after grouping and processing:", breakdown_with_totals.columns)

                # Ensure 'Total Net' is one of the columns
                if 'Total Net' in breakdown_with_totals.columns and 'Account Code Per Business' in breakdown_with_totals.columns:
                    # Create a summary table for account codes and their totals (Net column only)
                    summary_df = breakdown_with_totals[['Account Code Per Business', 'Total Net']].dropna(subset=['Total Net']).reset_index(drop=True)
                else:
                    # If the columns are not present, raise an error
                    raise ValueError("Required columns 'Account Code Per Business' or 'Total Net' are missing from the DataFrame")

                # Align the lengths of both DataFrames for concatenation
                max_len = max(len(breakdown_with_totals), len(summary_df))
                breakdown_with_totals.reset_index(drop=True, inplace=True)
                summary_df.reset_index(drop=True, inplace=True)

                # Align lengths for concatenation
                breakdown_with_totals = pd.concat([breakdown_with_totals, pd.DataFrame(index=range(max_len))], axis=1)
                summary_df = pd.concat([summary_df, pd.DataFrame(index=range(max_len))], axis=1)

                # Concatenate the breakdown with totals and the summary table side by side
                final_company_df = pd.concat([breakdown_with_totals, summary_df], axis=1)

                # Create the filename using the company name and code
                company_filename = f"Breakdown for {company_name_str} ({company_code}) - {selected_month_year_str_filenames}.csv"

                # Write the CSV content to a string
                csv_content = final_company_df.to_csv(index=False)

                # Write this string to the zip file
                breakdown_zip.writestr(company_filename, csv_content)

        breakdown_invoices_zip_stream.seek(0)
        zip_data2 = breakdown_invoices_zip_stream.read()

        # Convert company and tracking codes to strings
        final_df['Company Code'] = final_df['Company Code'].astype(str)
        final_df['Account Code Per Business'] = final_df['Account Code Per Business'].astype(str)
        final_df['Tracking Code'] = final_df['Tracking Code'].astype(str)
        final_df['Account Code'] = final_df['Account Code'].astype(str)

        # Join final_df with companies_df to include company names
        final_df = final_df.merge(companies_df, left_on='Company Code', right_on='company_code', how='left')
        final_df.drop(columns=['company_code'], inplace=True)

        # Prepare the data for Full Data Breakdown using 'Account Code with Descriptor' column
        final_df['Account Code with Descriptor'] = final_df['Account Code Per Business'] + ' - ' + final_df['Account Code Per Business Description']

        # Load the Excel template
        #template_path = '/Users/nyalpatel/Desktop/XeroAutomationWebApp/Recharging_Report_Template.xlsx'  # Replace this with the actual path to your template
        template_path = os.path.join(current_app.root_path, 'static', 'templates', 'Recharging_Report_Template.xlsx')

        workbook = load_workbook(template_path)
        
        # Select the "Full Data Breakdown" sheet
        sheet = workbook['Full Data Breakdown']  # Ensure the sheet name matches exactly

        # Write final_df data to the "Full Data Breakdown" sheet, starting from cell A3
        for row_idx, row in final_df.iterrows():
            for col_idx, value in enumerate(row):
                sheet.cell(row=row_idx + 4, column=col_idx + 1, value=value)  # Start writing at A3

        # Save the modified workbook to a BytesIO stream
        excel_stream = BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)  # Reset stream position




        # 3. Create combined_invoices.zip in memory
        combined_zip_stream = BytesIO()
        with zipfile.ZipFile(combined_zip_stream, 'w', zipfile.ZIP_DEFLATED) as combined_zip:
            # Add sales_invoice CSV
            sales_invoice_csv_filename = f"sales_invoice_for_{selected_month_year_str_filenames}.csv"
            combined_zip.writestr(sales_invoice_csv_filename, sales_invoice_df.to_csv(index=False))
            
            # Add purchase_invoices.zip
            combined_zip.writestr('purchase_invoices.zip', zip_data1)
            
            # Add breakdown_invoices.zip
            combined_zip.writestr('breakdown_invoices.zip', zip_data2)

            # Add the Excel report to the ZIP
            combined_zip.writestr('store_company_report.xlsx', excel_stream.read())
        
        combined_zip_stream.seek(0)

  

        # Define the base directory for temporary files within the app
        BASE_DIR = os.path.join(current_app.root_path, 'static', 'temp_files')

        # Ensure the directory exists
        os.makedirs(BASE_DIR, exist_ok=True)

        # Define the file path
        file_name = f"user_{user.username}_combined_invoices.zip"
        file_path = os.path.join(BASE_DIR, file_name)

        # Save the ZIP file to disk
        with open(file_path, "wb") as f:
            f.write(combined_zip_stream.read())

        # Mark task as completed
        if task_status:
            task_status.status = 'completed'
            task_status.result = "Task completed successfully."
            db.session.commit()

        return {"status": "success", "file_path": file_path}

    except Exception as e:
        add_log(f"Error occurred: {str(e)}", log_type="error", user_id=user_id)

        # Update the task status to failed
        if task_status:
            task_status.status = 'failed'
            task_status.result = str(e)
            db.session.commit()

        return {"status": "error", "message": str(e)}
    



def convert_parentheses_to_negative(value):
    # Remove commas
    value = value.replace(',', '')
    # Handle values in parentheses
    if re.match(r'^\(.*\)$', value):
        value = '-' + value[1:-1]
    return value


def combine_last_two_columns(df):
    # Define the possible column name pairs
    column_pairs = [
        ('Southern', 'Northern'),
        ('DOMINOS', 'ALL-GDK-COSTA-GYM'),
        ('tracking_group1', 'tracking_group2'), 
        ('DOMINOS', 'OTHER')
    ]
    
    # Iterate through the possible column name pairs
    for col1, col2 in column_pairs:
        if col1 in df.columns and col2 in df.columns:
            # Combine the two columns into 'Tracking Code'
            df['Tracking Code'] = df.apply(
                lambda row: row[col1] if pd.notna(row[col1]) and row[col1] != '' else row[col2],
                axis=1
            )
            # Drop the original columns
            df = df.drop(columns=[col1, col2])
            break  # Stop once we've handled one pair of columns
    
    return df



@shared_task(bind=True)
def update_invoice_record_task(self, user_id):
    try:
        # Step 1: Retrieve user-specific Xero credentials
        user = User.query.get(user_id)

        status = update_invoice_records(user)

        print(status)
    
    except Exception as e:
        self.update_state(state='FAILURE', meta={'exc': str(e)})
        raise e
    
import json

@shared_task(bind=True)
def refresh_xero_token_test(self, user_id):
    try:
        # Step 1: Retrieve user-specific Xero credentials
        user = User.query.get(user_id)
        token = json.loads(user.xero_token)
        refresh_token = token.get('refresh_token')

        refresh_xero_token(refresh_token, user)
    
    except Exception as e:
        self.update_state(state='FAILURE', meta={'exc': str(e)})
        raise e